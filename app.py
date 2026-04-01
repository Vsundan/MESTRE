from __future__ import annotations

import os
import sys
import uuid
from datetime import date, datetime, timedelta
import streamlit as st
import fitz
import anthropic
import json
import re
import time
import traceback
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from dotenv import dotenv_values
try:
    from pdfalign import align as pdfalign_align
    HAS_PDFALIGN = True
except ImportError:
    HAS_PDFALIGN = False

# Load API key — Railway env var takes priority, fall back to local claudbot .env
_ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
if not _ANTHROPIC_API_KEY:
    _env = dotenv_values(os.path.expanduser("~/claudbot/.env"))
    _ANTHROPIC_API_KEY = _env.get("ANTHROPIC_API_KEY_2") or _env.get("ANTHROPIC_API_KEY")

st.set_page_config(page_title="MESTRE", layout="wide")
st.title("MESTRE — Spec-Text Takeoff Engine")

# ─────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────
OPSS_NOTES = {
    "100": "General Conditions of Contract",
    "120": "Excavation — rock excavation, classification, measurement",
    "127": "Trenchless Installation of Pipes",
    "180": "Coffer Dams — design, installation, removal",
    "182": "Environmental Protection — spill prevention, species at risk, water quality",
    "206": "Grading — excavation tolerances, compaction, subgrade",
    "310": "Hot Mix Asphalt — placement, compaction, tack coat requirements",
    "314": "Granular Base/Subbase — gradation, lift thickness, compaction",
    "350": "Concrete Structures — forming, reinforcing, placement, curing",
    "351": "Concrete Sidewalks — base preparation, placement, finishing",
    "353": "Concrete Curbs — forming, placement, joints",
    "401": "Trenching — bedding classes, trench width, backfill requirements",
    "405": "Pipe Subdrains — installation, filter, outlet",
    "407": "Maintenance Holes and Catchbasins — precast, adjustment, frames",
    "410": "Storm/Sanitary Sewers — pipe installation, bedding, testing",
    "421": "Pipe Culverts — installation, end treatment, bedding",
    "441": "Watermain — installation, disinfection, pressure testing",
    "442": "Cathodic Protection — anodes, test stations, connections",
    "491": "Temporary Flow Control During Sewer and Watermain Construction",
    "493": "Temporary Water Supply — bypass, distribution, testing",
    "501": "Compacting — density requirements, testing frequency",
    "510": "Removals — existing structures, pavement, pipe",
    "517": "Dewatering — pumping, disposal, monitoring",
    "615": "Fencing — posts, fabric, installation",
    "620": "Electrical — conduit, wiring, connections",
    "706": "Traffic Control — signing, barriers, flagging, TCP",
    "802": "Topsoil — depth, placement, grading",
    "804": "Seeding — seed mix, fertilizer, maintenance",
    "805": "Erosion Control — silt fence, check dams, sediment basins",
    "902": "Excavating Structures — footings, backfill, frost tapers",
    "904": "Landscaping — sodding, mulch, maintenance",
    "928": "Access and Scaffolding — design, installation, safety",
    "930": "Concrete Surface Repair — patching, overlay",
    "1004": "Aggregates — clear stone gradation",
    "1010": "Aggregates — Granular A/B gradation, quality requirements",
    "1101": "Testing — compaction, gradation, concrete strength",
    "1150": "Pipe Materials — PVC, HDPE, concrete pipe specs",
    "1350": "Concrete Materials — mix design, air entrainment, strength",
    "1840": "Geotextiles — types, strength, installation",
    "1860": "Geotextiles — type, class, filtration requirements",
}

HISTORY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tender_history.json")


def get_hardcoded_opss_notes(opss_numbers: list) -> dict:
    """Return subset of OPSS_NOTES dict for the given spec numbers."""
    return {num: OPSS_NOTES.get(num, f"OPSS {num} — see spec document for details")
            for num in opss_numbers}


def get_opss_notes_from_db(opss_numbers: list) -> dict:
    """Query ChromaDB for real OPSS spec content; fall back to hardcoded dict."""
    chroma_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chromadb-store")
    notes = {}

    if not os.path.exists(chroma_path):
        return get_hardcoded_opss_notes(opss_numbers)

    try:
        import chromadb
        chroma_client = chromadb.PersistentClient(path=chroma_path)
        collection = chroma_client.get_collection("opss_specs")
        for num in opss_numbers:
            results = collection.query(
                query_texts=[f"OPSS {num} key requirements scope materials"],
                n_results=3,
                where={"opss_number": str(num).lstrip("0")},
            )
            if results["documents"][0]:
                combined = " ".join(results["documents"][0])
                notes[num] = combined[:200].rsplit(" ", 1)[0] + "..."
    except Exception:
        traceback.print_exc(file=sys.stderr)
        notes = {}

    # Fill gaps with hardcoded fallback
    hardcoded = get_hardcoded_opss_notes(opss_numbers)
    for num in opss_numbers:
        if num not in notes or not notes[num]:
            notes[num] = hardcoded.get(num, f"OPSS {num} — see spec document for details")

    return notes


CATEGORIES = {
    "Earthwork": ["excavat", "grading", "earth", "borrow", "backfill", "stockpile",
                  "capping", "clay cap", "import clay", "geogrid"],
    "Granular": ["granular", "base", "subbase", "aggregate", "select subgrade", "clear stone"],
    # NOTE: Asphalt keywords must stay broad but priority rules in categorize_item
    # override the generic CATEGORIES loop for items that should NOT be Asphalt
    # (e.g. "paving" in driveway/boulevard context is handled explicitly).
    "Asphalt": ["asphalt", "hot mix", "hma", "superpave", "tack coat", "milling",
                "paving", "boulevard pav", "driveway pav"],
    "Concrete": [
        "concrete", "formwork", "rebar", "reinforc", "curing",
        "sidewalk", "concrete sidewalk", "concrete slab", "concrete repair",
        "concrete restoration", "mudjacking", "mud jacking", "mud-jacking",
        "curb and gutter",
    ],
    # FIX 2: Structural/Masonry — bridge rehab, heritage, masonry repair scope
    "Structural": [
        "masonry", "repointing", "mortar", "stone resetting", "stone repair",
        "heritage", "bridge deck", "expansion joint", "bearing",
        "waterproofing membrane", "parapet", "barrier wall", "bridge jacking",
        "abutment", "wing wall", "retaining wall", "scaffolding",
        "work platform", "scaffold", "concrete patching",
    ],
    "Pipe/Sewer": ["sewer", "pipe", "culvert", "manhole", "maintenance hole", "catchbasin", "catch basin",
                   "drainage", "hdpe", "pvc", "cctv", "leachate",
                   "break into structure", "sanitary service", "cathodic protection"],
    "Watermain": ["watermain", "water main", "hydrant", "valve", "water service", "curb stop", "water service connection"],
    "Electrical": ["electrical", "conduit", "wiring", "lighting", "signal"],
    "Erosion Control": ["erosion", "silt", "sediment", "geotextile",
                        "straw bale", "check dam", "flow check"],
    "Landscaping": ["topsoil", "seed", "sod", "restoration", "landscap"],
    "Traffic": ["traffic", "sign", "barricade", "delineator"],
    "Demolition": ["remov", "demolit", "strip"],
    "Fencing": ["fence", "fencing", "gate", "litter fence"],
    "Equipment/Labour": ["hourly rate", "haulage", "equipment", "labour", "operator"],
    "General": [],
}

SCHEDULE_KEYWORDS = [
    "schedule of prices", "form of tender", "estimated quantity",
    "unit price", "tender item", "lump sum",
    "hourly rate", "haulage", "equipment", "labour", "subtotal",
    "total bid", "contractor's total", "tender price", "bid price",
    "item no", "spec no", "est. qty", "est qty",
    "provisional item", "contingency",
    "schedule of additional", "additional unit prices",
]

# Pages containing these phrases are spec/provisions sections — never schedule pages
EXCLUSION_KEYWORDS = [
    "general special provisions",
    "item special provisions",
    "information for tenderers",
    "general conditions",
    "standard contract forms",
    "supplemental general conditions",
]

CHECKLIST_CATEGORIES = ["Form", "Insurance", "Bonding", "WSIB", "Certificate",
                         "Schedule", "Document", "Submission Requirement", "Other"]
TIMELINE_FLAGS = ["DEADLINE", "MILESTONE", "PENALTY", "MEETING", "INFO"]

CHUNK_SIZE          = 100_000
CHUNK_OVERLAP       = 5_000
MAX_SCHEDULE_CHARS  = 160_000
FRONT_MATTER_CHARS  = 60_000   # used for checklist / timeline calls
HEADER_CHARS        = 10_000   # used for tender header extraction
CLAUDE_MODEL        = "claude-sonnet-4-20250514"
COST_PER_M_INPUT    = 3.0
CHARS_PER_TOKEN     = 4

HEADER_FILL   = PatternFill(start_color="B8860B", end_color="B8860B", fill_type="solid")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
PROV_FILL     = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LOW_CONF_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
RISK_HIGH     = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
RISK_MED      = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
RISK_LOW      = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
SECTION_FILL  = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SECTION_FONT  = Font(bold=True, color="FFFFFF", size=11)

# ─────────────────────────────────────────────
# Helper functions — existing
# ─────────────────────────────────────────────

def categorize_item(description: str, unit: str = "") -> str:
    desc_lower = str(description).lower()
    unit_upper = str(unit).upper()

    # Hourly rate items → Equipment/Labour regardless of description
    if unit_upper == "HOURS":
        return "Equipment/Labour"

    # ── Priority overrides (checked before generic keyword loop) ────────────────
    # ORDER MATTERS: non-structural overrides run first so that e.g. a manhole item
    # with "concrete" in the description does not get mis-tagged as Structural.

    # Pipe/Sewer priority rules: these items contain excavation/earthwork words in their
    # descriptions but are fundamentally pipe/sewer scope items.
    _pipe_priority_phrases = (
        "sanitary sewer", "storm sewer", "sanitary pipe", "pvc pipe", "hdpe pipe",
        "manhole", "maintenance hole", "catchbasin", "catch basin", "break into structure",
        "sanitary service", "sanitary lateral", "cathodic protection",
        "cctv", "leachate pipe", "culvert pipe",
    )
    if any(ph in desc_lower for ph in _pipe_priority_phrases):
        return "Pipe/Sewer"

    # Watermain priority: water service connections, curb stops, hydrants
    _watermain_priority_phrases = (
        "watermain", "water main", "water service connection", "curb stop",
        "hydrant", "gate valve",
    )
    if any(ph in desc_lower for ph in _watermain_priority_phrases):
        return "Watermain"

    # Asphalt priority: superpave, tack coat, HMA, driveways/boulevards in paving context
    _asphalt_priority_phrases = (
        "superpave", "tack coat", "hot mix asphalt", "hma", "milling",
        "asphalt binder", "asphalt surface", "cold in-place",
        "driveway restoration", "boulevard pav", "driveway pav",
        "asphalt pav", "paving asphalt",
    )
    if any(ph in desc_lower for ph in _asphalt_priority_phrases):
        return "Asphalt"

    # Structural/Masonry priority — masonry, bridge rehab, heritage, scaffolding.
    # Runs AFTER pipe/sewer and asphalt overrides so those always win for their items.
    # FIX 1: Expanded triggers per project requirements.
    _structural_priority_phrases = (
        # Masonry / stone repair
        "masonry", "repointing", "mortar", "stone resetting", "stone repair",
        "stone masonry", "resetting",
        # Bridge structural elements
        "bridge deck", "bridge rehabilit", "bridge repair",
        "expansion joint", "bearing pad", "elastomeric bearing",
        "abutment", "pier cap", "bridge pier",
        "wingwall", "wing wall", "soffit",
        "parapet", "barrier wall", "bridge jacking",
        # General structural
        "waterproofing membrane", "heritage",
        "structural steel", "beam", "girder",
        "rebar", "reinforcing steel", "reinforcement steel",
        "formwork", "scaffolding", "work platform",
        "access to work area", "access scaffolding",
        "concrete patching", "opss 928",
    )
    if any(ph in desc_lower for ph in _structural_priority_phrases):
        # Confirm the item is not an asphalt/road/pipe surface item
        _non_structural_kws = (
            "curb", "sidewalk", "boulevard", "asphalt", "paving", "granular",
            "manhole", "catchbasin", "sewer", "watermain",
        )
        _is_non_structural = any(kw in desc_lower for kw in _non_structural_kws)
        # Masonry/bridge/scaffolding terms win even over non-structural keywords
        # when the primary subject is clearly structural
        _strong_structural = any(ph in desc_lower for ph in (
            "masonry", "repointing", "mortar", "stone resetting", "stone repair",
            "stone masonry", "resetting", "abutment", "parapet", "soffit",
            "wingwall", "wing wall", "scaffolding", "access to work area",
            "bridge deck", "bridge rehabilit", "opss 928",
        ))
        if _strong_structural or not _is_non_structural:
            return "Structural"

    # "inspection" only qualifies as Pipe/Sewer when near pipe/leachate context
    if "inspection" in desc_lower and any(kw in desc_lower for kw in ("pipe", "leachate", "sewer", "culvert")):
        return "Pipe/Sewer"

    # Concrete maintenance scope: only treat lifting/levelling as concrete when
    # the same description clearly refers to sidewalk/concrete/slab/curb work.
    if any(kw in desc_lower for kw in ("lifting", "levelling", "leveling")):
        if any(ctx in desc_lower for ctx in ("sidewalk", "concrete", "slab", "curb and gutter")):
            return "Concrete"

    for cat, keywords in CATEGORIES.items():
        if cat == "General":
            continue
        if any(kw in desc_lower for kw in keywords):
            return cat
    return "General"


def extract_opss_refs(items: list) -> list:
    found = set()
    opss_explicit = re.compile(r"OPSS(?:\.PROV|\.MUNI)?\s+(\d{3,4})", re.IGNORECASE)
    sp_prefix = re.compile(r"(?:SP|Spec)\s+(\d{3,4})", re.IGNORECASE)
    bare_nums = re.compile(r"\b(\d{3,4})\b")
    for item in items:
        spec = str(item.get("spec_ref") or "")
        desc = str(item.get("description") or "")
        combined = spec + " " + desc
        for m in opss_explicit.finditer(combined):
            found.add(m.group(1))
        for m in sp_prefix.finditer(spec):
            found.add(m.group(1))
        for m in bare_nums.finditer(spec):
            if m.group(1) in OPSS_NOTES:
                found.add(m.group(1))
    return sorted(found, key=lambda x: int(x))


def extract_opss_from_full_text(full_text: str) -> list:
    """
    FIX 2 Pass 2: Regex scan of full document text for OPSS references.
    Finds patterns like "OPSS 706", "OPSS.MUNI 928", standalone codes near
    OPSS sections, and returns a sorted list of unique code strings.
    This supplements the Claude-based full scan (Pass 1) with zero API cost.
    """
    found = set()
    opss_explicit = re.compile(r"OPSS(?:\.PROV|\.MUNI)?\s*(\d{3,4})", re.IGNORECASE)
    for m in opss_explicit.finditer(full_text):
        code = m.group(1).lstrip("0") or "0"
        if code in OPSS_NOTES:
            found.add(code)
    # Also scan lines that look like a table of OPSS specs:
    # e.g. a line containing only a 3-4 digit number near a word like "November" or "April"
    # (typical OPSS reference tables list code, title, and revision date)
    _opss_table_line = re.compile(
        r"(?:^|\n)\s*(\d{3,4})\s+.{0,80}(?:January|February|March|April|May|June|July|August|"
        r"September|October|November|December)\s+\d{4}",
        re.IGNORECASE,
    )
    for m in _opss_table_line.finditer(full_text):
        code = m.group(1).lstrip("0") or "0"
        if code in OPSS_NOTES:
            found.add(code)
    return sorted(found, key=lambda x: int(x) if x.isdigit() else 9999)


def extract_other_standards_from_full_text(full_text: str) -> list[dict]:
    """
    Capture non-OPSS standards references for tenders that cite compliance
    standards but do not include OPSS codes.
    """
    patterns = [
        (
            "OTM Book 7",
            r"(?:Ontario Traffic Manual\s*\(\s*Book\s*7\s*\)|Ontario Traffic Manual.{0,40}Book\s*7|OTM\s*Book\s*7)",
            "Ontario Traffic Manual — Temporary Conditions (traffic control for construction zones)",
        ),
        (
            "OHSA",
            r"(?:Occupational Health\s*&?\s*Safety Act)",
            "Occupational Health & Safety Act — workplace safety compliance",
        ),
        (
            "WHMIS",
            r"\bWHMIS\b",
            "Workplace Hazardous Materials Information System — labeling and SDS compliance",
        ),
        (
            "AODA",
            r"(?:Accessibility for Ontarians with Disabilities)",
            "Accessibility for Ontarians with Disabilities Act",
        ),
        (
            "CSA",
            r"(?:Canadian Standard Association|Canadian Standards Association|\bCSA\b)",
            "Canadian Standards Association — product and equipment approvals",
        ),
        (
            "CCDC 2",
            r"\bCCDC\s*2\b",
            "Canadian Construction Documents Committee — Stipulated Price Contract",
        ),
        (
            "NBC",
            r"(?:National Building Code|\bNBC\b)",
            "National Building Code of Canada",
        ),
        (
            "OBC",
            r"(?:Ontario Building Code|\bOBC\b)",
            "Ontario Building Code",
        ),
    ]
    found = []
    seen = set()
    for code, pattern, description in patterns:
        if re.search(pattern, full_text, re.IGNORECASE):
            if code not in seen:
                seen.add(code)
                found.append({"code": code, "description": description})
    return found


def build_checklist_source_text(full_text: str) -> str:
    """
    Front matter usually contains standard bid forms, but custom submission
    requirements often live in scope/specification sections. Keep the prompt
    focused by appending the earliest spec block when available.
    """
    if len(full_text) <= FRONT_MATTER_CHARS:
        return full_text

    front = full_text[:FRONT_MATTER_CHARS]
    spec_start = None
    for marker in ("SPECIFICATIONS", "SCOPE OF WORK", "DETAIL", "SPECIAL PROVISIONS"):
        idx = full_text.upper().find(marker)
        if idx != -1 and idx > 0:
            if spec_start is None or idx < spec_start:
                spec_start = idx

    if spec_start is None:
        return front

    spec_chunk = full_text[spec_start: spec_start + (FRONT_MATTER_CHARS // 2)]
    if spec_chunk and spec_chunk not in front:
        return front + "\n\n" + spec_chunk
    return front


def _extract_sentence_like_excerpt(full_text: str, pattern: str, context_chars: int = 180) -> str:
    match = re.search(pattern, full_text, re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    start = max(0, match.start() - context_chars)
    end = min(len(full_text), match.end() + context_chars)
    excerpt = re.sub(r"\s+", " ", full_text[start:end]).strip()
    return excerpt[:220]


def _parse_irrevocable_days(full_text: str, timeline_items: list) -> int | None:
    for item in timeline_items:
        haystack = " ".join(str(item.get(field) or "") for field in ("event", "date", "risk_note"))
        m = re.search(r"(\d+)\s*day", haystack, re.IGNORECASE)
        if "irrevoc" in haystack.lower() and m:
            return int(m.group(1))

    text_lower = full_text.lower()
    word_map = {
        "thirty": 30,
        "sixty": 60,
        "ninety": 90,
        "one hundred and twenty": 120,
        "one-hundred-and-twenty": 120,
    }
    for match in re.finditer(r"irrevoc", text_lower):
        window = text_lower[match.start(): match.start() + 260]
        digit_match = re.search(r"\(?\b(\d{1,3})\b\)?\s*day", window)
        if digit_match:
            return int(digit_match.group(1))
        for words, value in word_map.items():
            if re.search(rf"{re.escape(words)}\s*(?:\(\s*\d+\s*\))?\s*day", window):
                return value
    return None


def check_if_item_is_bundled(item_keywords: tuple[str, ...], full_text: str) -> tuple[bool, str]:
    """
    Check whether a missing-scope item is explicitly bundled into contract price,
    assigned to the owner, or stated as not required.
    """
    if not full_text:
        return (False, "")

    lines = [re.sub(r"\s+", " ", line).strip() for line in full_text.splitlines()]
    lines = [line for line in lines if line]

    included_markers = (
        "included in the contract price",
        "included in the price",
        "shall be included",
        "included in contract price",
        "no extra cost",
        "no additional cost",
    )
    owner_markers = (
        "responsibility of the city",
        "responsibility of the owner",
        "provided by the city",
        "provided by the owner",
        "will be the responsibility of the city",
        "will be the responsibility of the owner",
    )
    not_required_markers = ("not required", "not applicable")

    for idx, line in enumerate(lines):
        window = " ".join(lines[max(0, idx - 3): min(len(lines), idx + 4)]).lower()
        if not any(keyword in window for keyword in item_keywords):
            continue
        if any(marker in window for marker in included_markers):
            return (True, "Included in contract price per tender specifications")
        if any(marker in window for marker in owner_markers):
            return (True, "Owner/City responsibility per tender specifications")
        if any(marker in window for marker in not_required_markers):
            return (True, "Not required per tender specifications")

    return (False, "")


_NON_SCHEDULE_ITEM_NO_VALUES = {"", "nan", "none", "null"}


def _normalize_item_no(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    raw = re.sub(r"\s+", "", raw)
    if re.fullmatch(r"\d+[A-Fa-f]", raw):
        return raw[:-1] + raw[-1].upper()
    return raw


def _is_valid_schedule_item_no(item_no: str) -> bool:
    raw = _normalize_item_no(item_no)
    if raw.lower() in _NON_SCHEDULE_ITEM_NO_VALUES:
        return False
    return bool(re.fullmatch(
        r"(?:"
        r"\d+(?:\.\d+)?[A-Za-z]?|"
        r"(?:FA|EQ)-?\d+|"
        r"[PE]-?\d+|"
        r"(?:LABOUR|LABOR)-\d+|"
        r"L-?\d+|"
        r"(?:EQUIPMENT)-\d+|"
        r"E-?\d+"
        r")",
        raw,
        re.IGNORECASE,
    ))


def _looks_like_rate_schedule_row(description: str, unit: str) -> tuple[bool, str]:
    desc_lower = str(description or "").strip().lower()
    unit_lower = str(unit or "").strip().lower()
    labour_markers = (
        "foreman",
        "tradesman",
        "skilled labour",
        "laborer",
        "labourer",
        "flag person",
        "flag persons",
        "heavy equipment operators",
        "equipment operator",
    )
    if any(marker in desc_lower for marker in labour_markers):
        return True, "LABOUR"
    if unit_lower in {"hours", "hour", "hr", "hrs"} and any(
        marker in desc_lower for marker in ("labour", "labor", "foreman", "tradesman", "flag person")
    ):
        return True, "LABOUR"
    return False, ""


def _coerce_quantity(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace(",", "")
    if not raw:
        return None
    if raw.endswith("%"):
        raw = raw[:-1].strip()
    try:
        return float(raw)
    except ValueError:
        return None


def _is_lump_sum_unit(unit: str) -> bool:
    normalized = re.sub(r"[^A-Z]", "", str(unit or "").upper())
    return normalized in {"LS", "LUMPSUM", "LUMP"}


def _is_percent_unit(unit: str) -> bool:
    normalized = re.sub(r"[^A-Z]", "", str(unit or "").upper())
    return normalized in {"PERCENT", "PCT"} or str(unit or "").strip() == "%"


def dedup_key(item: dict) -> tuple[str, str, str]:
    item_no = _normalize_item_no(item.get("item_no"))
    desc = str(item.get("description") or "").strip()
    qty = item.get("quantity")
    qty_value = _coerce_quantity(qty)
    if qty is None:
        qty_key = ""
    elif qty_value is None:
        qty_key = str(qty).strip()
    elif abs(qty_value - round(qty_value)) <= 1e-9:
        qty_key = str(int(round(qty_value)))
    else:
        qty_key = str(qty_value)
    return (item_no, desc, qty_key)


def validate_extraction(items: list) -> tuple[list, list]:
    warnings = []
    cleaned = []
    seen_exact = set()
    seen_item_nos: set = set()  # item_nos already in cleaned; used for spec-text dedup
    # fuzzy_seen: (item_no, desc[:30]) → index in cleaned; used for dedup keeping the qty version
    fuzzy_seen: dict = {}

    for idx, item in enumerate(items):
        label = item.get("item_no") or f"row {idx + 1}"
        for field in ("item_no", "spec_ref", "description", "unit"):
            val = item.get(field)
            if isinstance(val, str):
                item[field] = val.strip()
        item["item_no"] = _normalize_item_no(item.get("item_no"))
        label = item.get("item_no") or f"row {idx + 1}"
        desc = item.get("description") or ""
        if not desc:
            warnings.append(f"Item {label}: missing description — skipped")
            continue

        qty = item.get("quantity")
        if qty is not None:
            try:
                item["quantity"] = float(qty)
            except (TypeError, ValueError):
                warnings.append(f"Item {label} ({desc[:40]}): non-numeric quantity '{qty}' — set to null")
                item["quantity"] = None
                qty = None

        unit = item.get("unit") or ""
        if qty is not None and not unit:
            item["unit"] = "Missing unit"
            warnings.append(f"Item {label} ({desc[:40]}): quantity present but unit missing")
        if qty is None and not unit:
            ls_hint = any(kw in desc.lower() for kw in ["lump sum", " ls", "ls "])
            if not ls_hint:
                warnings.append(f"Item {label} ({desc[:40]}): no quantity or unit — flagged 'Check manually'")
                item["unit"] = "Check manually"
        conf = item.get("confidence")
        if conf is None:
            item["confidence"] = 0.5
        else:
            try:
                item["confidence"] = max(0.0, min(1.0, float(conf)))
            except (TypeError, ValueError):
                item["confidence"] = 0.5

        # Exact dedup: section-aware and quantity-aware.
        exact_key = dedup_key(item)
        if exact_key in seen_exact:
            warnings.append(f"Item {label} ({desc[:40]}): exact duplicate — skipped")
            continue
        seen_exact.add(exact_key)

        # Spec-text dedup: same item_no + no quantity + unit is "unit price" or "lump sum"
        # These are Item Special Provisions re-extractions, not real schedule rows
        item_no = str(item.get("item_no") or "")
        unit_lower = str(item.get("unit") or "").lower().strip()
        if (item_no and item_no in seen_item_nos
                and item.get("quantity") is None
                and unit_lower in ("unit price", "lump sum")):
            warnings.append(f"Item {label} ({desc[:40]}): spec-text duplicate (no qty, unit='{unit_lower}') — skipped")
            continue

        # Fuzzy dedup: same item_no + first 30 chars of description.
        # Keep different quantities as separate items; multi-section suffixes
        # should survive in item_no and must never be replaced with fake labels.
        fuzzy_key = (item_no, desc[:30].lower())
        if fuzzy_key in fuzzy_seen and item_no:  # only fuzzy-dedup when item_no is present
            existing_idx = fuzzy_seen[fuzzy_key]
            existing = cleaned[existing_idx]
            existing_qty = existing.get("quantity")
            new_qty = item.get("quantity")

            def _qtys_match(a, b) -> bool:
                if a is None and b is None:
                    return True
                if a is None or b is None:
                    return False
                try:
                    return abs(float(a) - float(b)) <= 1e-6
                except (TypeError, ValueError):
                    return str(a) == str(b)

            if existing_qty is None and new_qty is not None:
                # Same quantity context but existing had no qty — upgrade existing
                warnings.append(f"Item {label} ({desc[:30]}): fuzzy duplicate replaced — kept version with quantity")
                cleaned[existing_idx] = item
                continue
            elif _qtys_match(existing_qty, new_qty):
                warnings.append(f"Item {label} ({desc[:30]}): fuzzy duplicate — skipped")
                continue
            else:
                warnings.append(
                    f"Item {label} ({desc[:30]}): same item_no/description but different quantity "
                    f"({existing_qty} vs {new_qty}) — kept as separate item"
                )

        if item_no:
            seen_item_nos.add(item_no)
        fuzzy_seen[fuzzy_key] = len(cleaned)
        cleaned.append(item)
    return cleaned, warnings


def split_items_by_quality(items: list) -> tuple[list, list]:
    """
    Split items into confirmed takeoff items and possible additional items.
    Possible items: unit == "Check manually" (no qty, no unit, not lump sum).
    These go to a separate section — not the main Takeoff sheet.
    """
    main_items = []
    possible_items = []
    confirmed_item_nos = {
        _normalize_item_no(item.get("item_no"))
        for item in items
        if str(item.get("unit", "")).strip().lower() != "check manually"
    }
    confirmed_descs = {
        re.sub(r"\s+", " ", str(item.get("description") or "").strip().lower())
        for item in items
        if str(item.get("unit", "")).strip().lower() != "check manually"
    }
    for item in items:
        if str(item.get("unit", "")).strip().lower() == "check manually":
            item_no_key = _normalize_item_no(item.get("item_no"))
            desc_key = re.sub(r"\s+", " ", str(item.get("description") or "").strip().lower())
            if item_no_key and item_no_key in confirmed_item_nos:
                continue
            if desc_key and desc_key in confirmed_descs:
                continue
            possible_items.append(item)
        else:
            main_items.append(item)
    return main_items, possible_items


def extract_summary_rows(schedule_text: str) -> list:
    """
    FIX 4: Scan the bottom of the schedule text for summary/total rows:
    Tender Price (excluding HST), Contingency (%), HST (%), Total Tender Price.
    Returns a list of summary row dicts with category="Summary".
    These are appended to the Takeoff sheet but never counted in item statistics.
    """
    summary_rows = []
    lines = schedule_text.splitlines()

    # Patterns to detect summary rows (case-insensitive)
    _summary_patterns = [
        # Subtotal / Tender Price (excluding HST)
        (re.compile(r"tender\s+price\s*\(\s*excluding\s+hst\s*\)", re.IGNORECASE), "SUBTOTAL"),
        (re.compile(r"sub[-\s]?total", re.IGNORECASE), "SUBTOTAL"),
        (re.compile(r"subtotal", re.IGNORECASE), "SUBTOTAL"),
        # Contingency with %
        (re.compile(r"contingency\s*\(\s*\d+\s*%\s*\)", re.IGNORECASE), "CONTINGENCY"),
        (re.compile(r"contingency\s+\d+\s*%", re.IGNORECASE), "CONTINGENCY"),
        # HST / Harmonized Sales Tax
        (re.compile(r"hst\s*\(\s*\d+\s*%\s*\)", re.IGNORECASE), "TAX"),
        (re.compile(r"harmonized\s+sales\s+tax\s*\(\s*\d+\s*%\s*\)", re.IGNORECASE), "TAX"),
        # Total Tender Price / Grand Total
        (re.compile(r"total\s+tender\s+price\s*\(\s*including\s+hst\s*\)", re.IGNORECASE), "TOTAL"),
        (re.compile(r"total\s+tender\s+price", re.IGNORECASE), "TOTAL"),
        (re.compile(r"grand\s+total", re.IGNORECASE), "TOTAL"),
    ]

    seen_row_keys = set()
    for line in lines:
        line_stripped = line.strip()
        if not line_stripped:
            continue
        # Summary rows are short bid-form labels, not narrative sentences.
        if len(line_stripped) > 80:
            continue
        for pattern, row_type in _summary_patterns:
            if pattern.fullmatch(line_stripped):
                row_key = (row_type, re.sub(r"\s+", " ", line_stripped).lower()[:60])
                if row_key in seen_row_keys:
                    break
                seen_row_keys.add(row_key)
                summary_rows.append({
                    "item_no": "",
                    "spec_ref": "",
                    "description": line_stripped,
                    "quantity": None,
                    "unit": "",
                    "is_provisional": False,
                    "confidence": 1.0,
                    "category": "Summary",
                })
                break  # one match per line

    return summary_rows


def fix_lump_sum_quantities(items: list) -> tuple[list, list]:
    """Convert Ontario 100% lump-sum rows to quantity=1 before validation/XLSX."""
    warnings = []
    for item in items:
        label = _normalize_item_no(item.get("item_no")) or "row"
        desc_lower = str(item.get("description") or "").lower()
        qty_raw = item.get("quantity")
        qty_value = _coerce_quantity(qty_raw)
        unit = str(item.get("unit") or "").strip()

        if "hst" in desc_lower or "harmonized sales tax" in desc_lower:
            continue

        if _is_percent_unit(unit):
            item["quantity"] = 1.0
            item["unit"] = "LS"
            warnings.append(f"Item {label}: percentage unit '{unit}' corrected to qty=1 LS")
            continue

        if _is_lump_sum_unit(unit):
            qty_text = str(qty_raw or "").strip().replace(" ", "")
            if qty_text.endswith("%") or (qty_value is not None and abs(qty_value - 100.0) < 1e-6):
                item["quantity"] = 1.0
                warnings.append(f"Item {label}: lump sum quantity '{qty_raw}' corrected to 1")

    return items, warnings


def fix_hst_percentage_items(items: list) -> tuple[list, list]:
    """Normalize HST rows to percentage quantities instead of lump-sum placeholders."""
    warnings = []
    for item in items:
        description = str(item.get("description") or "")
        if not re.search(r"\b(?:hst|harmonized sales tax)\b", description, re.IGNORECASE):
            continue
        match = re.search(r"(\d+(?:\.\d+)?)\s*%", description)
        if not match:
            continue

        percent_value = float(match.group(1))
        if abs(percent_value - round(percent_value)) <= 1e-9:
            percent_value = int(round(percent_value))

        qty_key = _normalize_quantity_for_matching(item.get("quantity"))
        unit_key = _normalize_rate_schedule_unit_for_matching(item.get("unit"))
        if qty_key == percent_value and unit_key == "%":
            continue

        item["quantity"] = percent_value
        item["unit"] = "%"
        label = _normalize_item_no(item.get("item_no")) or "row"
        warnings.append(f"Item {label}: normalized HST row to quantity={percent_value} unit=%")

    return items, warnings


def filter_non_schedule_items(items: list) -> tuple[list, list]:
    """Remove non-schedule text and normalize/guard item identifiers."""
    filtered = []
    warnings = []
    synthetic_counts = {"LABOUR": 0, "EQUIPMENT": 0}

    for idx, item in enumerate(items):
        item_no = _normalize_item_no(item.get("item_no"))
        desc = str(item.get("description") or "").strip()
        unit = str(item.get("unit") or "").strip()

        if item_no.lower() in _NON_SCHEDULE_ITEM_NO_VALUES:
            looks_like_rate, prefix = _looks_like_rate_schedule_row(desc, unit)
            if looks_like_rate:
                synthetic_counts[prefix] += 1
                item_no = f"{prefix}-{synthetic_counts[prefix]}"
                item["item_no"] = item_no
                warnings.append(
                    f"Item row {idx + 1} ({desc[:50]}): assigned synthetic item_no '{item_no}' "
                    "for structured labour/equipment rate row"
                )
            else:
                warnings.append(
                    f"Item row {idx + 1} ({desc[:50]}): no valid item_no — excluded as non-schedule text"
                )
                continue
        else:
            item["item_no"] = item_no

        if not _is_valid_schedule_item_no(item_no):
            warnings.append(
                f"Item {item_no or f'row {idx + 1}'} ({desc[:50]}): invalid item_no format — excluded"
            )
            continue

        filtered.append(item)

    return filtered, warnings


def ensure_labour_rate_items(schedule_text: str, items: list) -> tuple[list, list]:
    """Backfill structured labour-rate rows when the table is present but Claude omits them."""
    warnings = []
    text_lower = schedule_text.lower()
    if "labour and equipment rates" not in text_lower or "additional labour requirements" not in text_lower:
        return items, warnings

    existing_labour_descs = {
        str(item.get("description") or "").strip().lower()
        for item in items
        if str(item.get("item_no") or "").upper().startswith("LABOUR-")
    }
    if existing_labour_descs:
        return items, warnings

    lines = [line.strip() for line in schedule_text.splitlines()]
    in_labour_section = False
    labour_candidates: list[str] = []
    idx = 0
    while idx < len(lines):
        line = lines[idx]
        lower = line.lower()
        if "additional labour requirements" in lower:
            in_labour_section = True
            idx += 1
            continue
        if in_labour_section and "additional equipment requirements" in lower:
            break
        if in_labour_section and line:
            if lower == "tradesman:" and idx + 1 < len(lines):
                next_line = lines[idx + 1].strip()
                if next_line:
                    labour_candidates.append(f"Tradesman: {next_line}")
                    idx += 2
                    continue
            if lower in {
                "foreman",
                "skilled labour",
                "labourer",
                "flag persons",
                "heavy equipment operators",
            }:
                labour_candidates.append(line)
        idx += 1

    for description in labour_candidates:
        desc_lower = description.lower()
        if desc_lower in existing_labour_descs:
            continue
        item_no = f"LABOUR-{len(existing_labour_descs) + 1}"
        items.append({
            "item_no": item_no,
            "spec_ref": "",
            "description": description,
            "quantity": None,
            "unit": "HOURS",
            "is_provisional": False,
            "confidence": 0.8,
        })
        existing_labour_descs.add(desc_lower)
        warnings.append(f"Added fallback labour rate row {item_no}: {description}")

    return items, warnings


def ensure_numbered_rate_items(schedule_text: str, items: list) -> tuple[list, list]:
    """Normalize numbered labour/equipment rate schedules to L#/E# item identifiers."""
    warnings = []
    marker = "SCHEDULE OF ADDITIONAL UNIT PRICES"
    if marker not in schedule_text.upper():
        return items, warnings

    schedule_upper = schedule_text.upper()
    start_idx = schedule_upper.find(marker)
    block = schedule_text[start_idx:]
    block_upper = block.upper()
    for end_marker in ("TOTAL ITEMS 1 TO 20", "PERIOD OF VALIDITY OF TENDER"):
        end_idx = block_upper.find(end_marker)
        if end_idx != -1:
            block = block[:end_idx]
            break

    lines = [line.strip() for line in block.splitlines()]
    current_group = ""
    parsed_items = []
    idx = 0
    while idx < len(lines):
        line = lines[idx]
        upper = line.upper()
        if upper == "LABOUR":
            current_group = "LABOUR"
            idx += 1
            continue
        if upper == "EQUIPMENT":
            current_group = "EQUIPMENT"
            idx += 1
            continue

        match = re.match(r"^(\d+)\.\s+(.*\S)$", line)
        if not match or current_group not in {"LABOUR", "EQUIPMENT"}:
            idx += 1
            continue

        number = int(match.group(1))
        description_parts = [match.group(2)]
        rate = None
        idx += 1
        while idx < len(lines):
            nxt = lines[idx]
            nxt_upper = nxt.upper()
            if not nxt:
                idx += 1
                continue
            if _looks_like_pdf_header_footer_line(nxt):
                break
            if re.match(r"^\d+\.\s+", nxt) or nxt_upper in {"LABOUR", "EQUIPMENT"}:
                break
            if nxt_upper in {
                "DESCRIPTION",
                "HOURLY RATE",
                "HOURS",
                "SUBTOTAL",
                "(BID PRICE)",
            }:
                idx += 1
                continue
            if re.fullmatch(r"\d+(?:\.\d+)?", nxt) and rate is None:
                rate = float(nxt)
                idx += 1
                continue
            if nxt == "$":
                idx += 1
                continue
            if re.fullmatch(r"[A-Za-z/&() -]*Equipment:?", nxt):
                idx += 1
                continue
            description_parts.append(nxt)
            idx += 1

        description = " ".join(part for part in description_parts if part).strip()
        description = _clean_rate_schedule_description(description)
        if not description:
            continue
        item_no = f"L{number}" if current_group == "LABOUR" else f"E{number}"
        parsed_items.append({
            "item_no": item_no,
            "spec_ref": "",
            "description": description,
            "quantity": rate,
            "unit": "HOURS",
            "is_provisional": False,
            "confidence": 0.8,
        })

    if not parsed_items:
        return items, warnings

    added_count = 0
    skipped_count = 0
    normalized_count = 0
    for parsed in parsed_items:
        match_idx = _find_matching_rate_schedule_item(parsed, items)
        if match_idx is not None:
            existing_desc = _clean_rate_schedule_description(items[match_idx].get("description"))
            if existing_desc and existing_desc != str(items[match_idx].get("description") or "").strip():
                items[match_idx]["description"] = existing_desc
                normalized_count += 1
            skipped_count += 1
            continue

        same_id_idx = next(
            (
                idx for idx, item in enumerate(items)
                if _normalize_item_no(item.get("item_no")) == parsed["item_no"]
            ),
            None,
        )
        if same_id_idx is not None:
            items[same_id_idx].update(parsed)
            normalized_count += 1
            continue

        items.append(parsed)
        added_count += 1

    if skipped_count:
        warnings.append(
            f"Skipped {skipped_count} numbered rate row(s) already captured by extraction"
        )
    if normalized_count:
        warnings.append(
            f"Normalized {normalized_count} numbered rate row(s) to cleaned schedule formatting"
        )
    if added_count:
        warnings.append(f"Added {added_count} numbered L#/E# rate row(s) from schedule text")

    return items, warnings

def _looks_like_pdf_header_footer_line(value: object) -> bool:
    line = re.sub(r"\s+", " ", str(value or "")).strip()
    if not line:
        return False
    lower = line.lower()
    if re.search(r"\bpage\s*\d+\b", lower) and re.search(
        r"\b(form of tender|contract\s*no\.?|division\s+\d+|project\s+no\.?|"
        r"authority|inc\.?|limited|ltd\.?|consulting|engineering|"
        r"\d{3}-\d{5}-\d{2}|"
        r"(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s+\d{4})\b",
        lower,
    ):
        return True
    if "form of tender" in lower and (
        "contract no" in lower or "division " in lower or re.search(r"\d{3}-\d{5}-\d{2}", lower)
    ):
        return True
    return False


def _clean_rate_schedule_description(value: object) -> str:
    """
    Remove header/footer fragments that can bleed into parsed rate-schedule
    descriptions when a table spans page breaks.
    """
    raw = str(value or "")
    if not raw.strip():
        return ""
    parts = []
    for line in raw.splitlines():
        cleaned_line = re.sub(r"\s+", " ", line).strip()
        if not cleaned_line or _looks_like_pdf_header_footer_line(cleaned_line):
            continue
        parts.append(cleaned_line)
    cleaned = " ".join(parts) if parts else re.sub(r"\s+", " ", raw).strip()

    tail_markers = (
        r"\bpage\s*\d+\b",
        r"\bform\s+of\s+tender\b",
        r"\bdivision\s+\d+\b",
        r"\bcontract\s*no\.?\b",
        r"\bproject\s+no\.?\b",
        r"\bwsp\s+canada\b",
        r"\d{3}-\d{5}-\d{2}",
    )
    earliest = None
    for pattern in tail_markers:
        match = re.search(pattern, cleaned, re.IGNORECASE)
        if match:
            earliest = match.start() if earliest is None else min(earliest, match.start())
    if earliest is not None:
        cleaned = cleaned[:earliest]

    cleaned = re.sub(
        r"\s+[A-Z][A-Za-z0-9&.,'’/\-]*(?:\s+[A-Z][A-Za-z0-9&.,'’/\-]*)*\s+"
        r"(?:Authority|Inc\.?|Ltd\.?|Limited|Corporation|Corp\.?|County|City|"
        r"Town(?:ship)?|Municipality|Region(?:al)?|Consultants?|Engineering|Canada)$",
        "",
        cleaned,
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -,:;/")
    return cleaned


def _normalize_rate_schedule_unit_for_matching(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    normalized = re.sub(r"[^A-Z%]", "", raw.upper())
    if normalized in {"HR", "HRS", "HOUR", "HOURS"}:
        return "hours"
    if normalized in {"LS", "LUMP", "LUMPSUM"}:
        return "ls"
    if normalized in {"PERCENT", "PCT"} or raw == "%":
        return "%"
    return raw.lower()


def _normalize_quantity_for_matching(value: object) -> object:
    qty_value = _coerce_quantity(value)
    if qty_value is None:
        raw = str(value or "").strip()
        return None if not raw else raw.lower()
    if abs(qty_value - round(qty_value)) <= 1e-9:
        return int(round(qty_value))
    return round(qty_value, 6)


def _rate_schedule_signature(description: object, quantity: object, unit: object) -> tuple[str, object, str]:
    cleaned_desc = _clean_rate_schedule_description(description).lower()
    cleaned_desc = re.sub(r"[^a-z0-9]+", " ", cleaned_desc)
    cleaned_desc = " ".join(cleaned_desc.split())
    return (
        cleaned_desc,
        _normalize_quantity_for_matching(quantity),
        _normalize_rate_schedule_unit_for_matching(unit),
    )


def _find_matching_rate_schedule_item(parsed_item: dict, items: list) -> int | None:
    parsed_sig = _rate_schedule_signature(
        parsed_item.get("description"),
        parsed_item.get("quantity"),
        parsed_item.get("unit"),
    )
    if not parsed_sig[0]:
        return None
    for idx, existing in enumerate(items):
        existing_sig = _rate_schedule_signature(
            existing.get("description"),
            existing.get("quantity"),
            existing.get("unit"),
        )
        if existing_sig == parsed_sig:
            return idx
    return None


def ensure_force_account_rate_items(schedule_text: str, items: list, full_text: str = "") -> tuple[list, list]:
    """
    Backfill "Schedule of Force Account Rates" rows when the model skips blank rate tables.
    Prefer schedule_text to avoid table-of-contents/narrative false positives, and fall back
    to full_text only if the schedule slice did not include the force account page.
    """
    warnings = []
    search_sources = [schedule_text]
    if full_text and full_text != schedule_text:
        search_sources.append(full_text)

    personnel_entries: list[str] = []
    equipment_entries: list[str] = []
    equipment_suffix = ""
    skip_exact = {
        "list by occupation",
        "hourly rate",
        "overtime hourly rate",
        "description",
        "model and size",
        "other (list)",
    }

    for source_text in search_sources:
        lines = [line.strip() for line in source_text.splitlines()]
        for idx, line in enumerate(lines):
            if "schedule of force account rates" not in line.lower():
                continue

            window = lines[idx:min(len(lines), idx + 120)]
            if not any(entry.lower() == "personnel" for entry in window):
                continue
            if not any(entry.lower().startswith("equipment") for entry in window):
                continue

            current_section = ""
            local_personnel: list[str] = []
            local_equipment: list[str] = []
            local_suffix = ""

            for candidate in window:
                lowered = candidate.lower()
                if not candidate:
                    continue
                if _looks_like_pdf_header_footer_line(candidate):
                    continue
                if lowered == "personnel":
                    current_section = "personnel"
                    continue
                if re.match(r"^equipment\s*(?::|\()", lowered) or lowered == "equipment":
                    current_section = "equipment"
                    if "complete with operator" in lowered:
                        local_suffix = " (COMPLETE WITH OPERATOR)"
                    continue
                if candidate.isdigit():
                    continue
                if lowered in skip_exact:
                    continue
                if current_section == "personnel":
                    if re.fullmatch(r"[A-Z][A-Z\s/&()\-]{8,}", candidate):
                        break
                    cleaned_candidate = _clean_rate_schedule_description(candidate)
                    if cleaned_candidate:
                        local_personnel.append(cleaned_candidate)
                    continue
                if current_section == "equipment":
                    if lowered == "other (list)":
                        current_section = ""
                        break
                    if re.fullmatch(r"[A-Z][A-Z\s/&()\-]{8,}", candidate):
                        break
                    cleaned_candidate = _clean_rate_schedule_description(candidate)
                    if cleaned_candidate:
                        local_equipment.append(cleaned_candidate)

            local_personnel = [entry for entry in local_personnel if entry]
            local_equipment = [entry for entry in local_equipment if entry]
            if local_personnel or local_equipment:
                personnel_entries = local_personnel
                equipment_entries = local_equipment
                equipment_suffix = local_suffix
                break

        if personnel_entries or equipment_entries:
            break

    if not personnel_entries and not equipment_entries:
        return items, warnings

    def _make_force_account_row(prefix: str, number: int, description: str, spec_ref: str) -> dict:
        normalized_desc = _clean_rate_schedule_description(description)
        if prefix == "EQ" and equipment_suffix and "complete with operator" not in description.lower():
            normalized_desc = f"{description}{equipment_suffix}"
        normalized_desc = _clean_rate_schedule_description(normalized_desc)
        return {
            "item_no": f"{prefix}-{number}",
            "spec_ref": spec_ref,
            "description": normalized_desc,
            "quantity": None,
            "unit": "HOURS",
            "is_provisional": False,
            "confidence": 1.0,
            "category": "Equipment/Labour",
        }

    parsed_rows = []
    for idx, description in enumerate(personnel_entries, 1):
        row = _make_force_account_row("FA", idx, description, "Force Account")
        if row["description"]:
            parsed_rows.append(row)
    for idx, description in enumerate(equipment_entries, 1):
        row = _make_force_account_row("EQ", idx, description, "Force Account Equipment")
        if row["description"]:
            parsed_rows.append(row)

    normalized_count = 0
    added_count = 0
    skipped_count = 0
    for parsed_row in parsed_rows:
        match_idx = _find_matching_rate_schedule_item(parsed_row, items)
        if match_idx is not None:
            existing = items[match_idx]
            existing.update({
                "item_no": parsed_row["item_no"],
                "spec_ref": parsed_row["spec_ref"],
                "description": parsed_row["description"],
                "unit": parsed_row["unit"],
                "is_provisional": False,
                "category": "Equipment/Labour",
            })
            existing["confidence"] = max(float(existing.get("confidence") or 0), parsed_row["confidence"])
            normalized_count += 1
            skipped_count += 1
            continue

        same_id_idx = next(
            (
                idx for idx, item in enumerate(items)
                if _normalize_item_no(item.get("item_no")) == parsed_row["item_no"]
            ),
            None,
        )
        if same_id_idx is not None:
            items[same_id_idx].update(parsed_row)
            normalized_count += 1
            continue

        items.append(parsed_row)
        added_count += 1

    if skipped_count:
        warnings.append(
            f"Skipped {skipped_count} force account row(s) already captured by extraction"
        )
    if normalized_count:
        warnings.append(
            f"Normalized {normalized_count} force account row(s) to deterministic FA/EQ formatting"
        )
    if added_count:
        warnings.append(f"Added {added_count} deterministic force account row(s) from schedule text")

    return items, warnings



def _page_is_excluded(text: str) -> bool:
    """
    Return True only if an exclusion keyword appears as a section heading
    (at the start of a line, possibly with leading whitespace).
    Mid-sentence references like "Section X of the General Conditions" are ignored.
    """
    import re as _re
    lower = text.lower()
    for excl in EXCLUSION_KEYWORDS:
        # Must appear at start of a line (after optional whitespace/bullets)
        if _re.search(r'(?:^|\n)\s*' + _re.escape(excl), lower):
            return True
    return False


def find_schedule_page_indices(pages_text: list) -> list:
    raw_hits = []
    for i, text in enumerate(pages_text):
        lower = text.lower()
        # Skip pages where exclusion keyword is a section heading (not a mid-sentence reference)
        if _page_is_excluded(text):
            continue
        hits = sum(1 for kw in SCHEDULE_KEYWORDS if kw in lower)
        if hits >= 2:
            raw_hits.append(i)
    if not raw_hits:
        return []
    # Bridge small gaps (< 5 pages) between detected schedule pages
    bridged = set(raw_hits)
    for i in range(len(raw_hits) - 1):
        a, b = raw_hits[i], raw_hits[i + 1]
        if b - a < 5:
            for mid in range(a + 1, b):
                bridged.add(mid)
    return sorted(bridged)


def _cluster_pages(detected: list, max_gap: int = 4) -> list[list]:
    """Group detected page indices into clusters where consecutive pages are within max_gap."""
    if not detected:
        return []
    clusters = [[detected[0]]]
    for i in range(1, len(detected)):
        if detected[i] - detected[i - 1] <= max_gap:
            clusters[-1].append(detected[i])
        else:
            clusters.append([detected[i]])
    return clusters


def _page_has_high_conf_schedule_markers(text: str) -> bool:
    lower = text.lower()
    has_schedule_header = (
        "schedule of items and prices" in lower
        or "schedule of prices" in lower
    )
    has_quantity_col = bool(
        re.search(r"(?:tender\s*quantity|est\.?\s*qty|(?:^|\s)qty(?:\s|$)|(?:^|\s)quantity(?:\s|$))", lower)
    )
    has_table_core = (
        bool(re.search(r"item\s*no", lower))
        and "description" in lower
        and has_quantity_col
        and "unit" in lower
    )
    has_pricing_cols = bool(re.search(r"unit\s*price", lower)) and "amount" in lower
    has_part_total = bool(re.search(r"total\s+part\s+[a-f]", lower))
    return bool(
        (has_schedule_header and (has_table_core or has_pricing_cols or has_part_total))
        or (has_table_core and has_pricing_cols)
        or (has_part_total and has_table_core)
    )


def _expand_page_context(page_indices: list[int], total_pages: int, radius: int = 1) -> list[int]:
    expanded = set()
    for idx in page_indices:
        start = max(0, idx - radius)
        end = min(total_pages - 1, idx + radius)
        expanded.update(range(start, end + 1))
    return sorted(expanded)


def _page_is_schedule_continuation(text: str) -> bool:
    lower = text.lower()
    marker_hits = sum(
        1
        for pattern in (
            r"item\s*no",
            r"description|(?:^|\s)item(?:\s|$)",
            r"tender\s*quantity|(?:^|\s)quantity(?:\s|$)",
            r"(?:^|\s)unit(?:\s|$)",
            r"unit\s*price",
            r"amount",
        )
        if re.search(pattern, lower)
    )
    has_rate_table = (
        "description" in lower
        and "hourly rate" in lower
        and "subtotal" in lower
        and "hours" in lower
    )
    return bool(
        marker_hits >= 4
        or has_rate_table
        or re.search(r"\bpart\s+[a-f]\b", lower)
        or "provisional" in lower
        or "labour and equipment rates" in lower
        or "schedule of additional unit prices" in lower
    )


def _page_has_rate_schedule_markers(text: str) -> bool:
    lower = text.lower()
    has_explicit_rate_header = (
        "schedule of additional unit prices" in lower
        or "schedule of force account rates" in lower
    )
    has_named_rate_section = (
        "labour and equipment rates" in lower
        and (
            "additional labour requirements" in lower
            or "additional equipment requirements" in lower
            or "price/hr" in lower
            or "price/day" in lower
        )
    )
    has_rate_table = (
        "description" in lower
        and "hourly rate" in lower
        and "subtotal" in lower
        and "hours" in lower
        and bool(re.search(r"(?:^|\n)\s*\d{1,2}[.)]\s", text))
    )
    return bool(
        has_explicit_rate_header
        or has_named_rate_section
        or has_rate_table
    )


def _collect_adjacent_detected_pages(detected: list[int], pages_text: list[str], seed_pages: list[int]) -> list[int]:
    detected_set = set(detected)
    included = set()

    def qualifies(idx: int) -> bool:
        text = pages_text[idx]
        return (
            _page_has_high_conf_schedule_markers(text)
            or _page_is_schedule_continuation(text)
            or _page_has_rate_schedule_markers(text)
        )

    for seed in seed_pages:
        if seed not in detected_set:
            continue
        included.add(seed)
        prev = seed - 1
        while prev in detected_set and qualifies(prev):
            included.add(prev)
            prev -= 1
        nxt = seed + 1
        while nxt in detected_set and qualifies(nxt):
            included.add(nxt)
            nxt += 1
    return sorted(included)


def _expected_multi_section_letters(schedule_text: str) -> set[str]:
    return {match.group(1).upper() for match in re.finditer(r"\bpart\s+([A-F])\b", schedule_text, re.IGNORECASE)}


def _extracted_multi_section_letters(items: list) -> set[str]:
    letters = set()
    for item in items:
        raw = _normalize_item_no(item.get("item_no"))
        match = re.fullmatch(r"\d+(?:\.\d+)?([A-F])", raw)
        if match:
            letters.add(match.group(1).upper())
    return letters


def build_schedule_text(pages_text: list, full_scan: bool) -> tuple[str, list]:
    if full_scan:
        return "\n\n".join(pages_text), list(range(len(pages_text)))
    detected = find_schedule_page_indices(pages_text)
    if not detected:
        return "\n\n".join(pages_text), list(range(len(pages_text)))
    high_conf_pages = [idx for idx in detected if _page_has_high_conf_schedule_markers(pages_text[idx])]
    if high_conf_pages:
        rate_pages = [idx for idx in detected if _page_has_rate_schedule_markers(pages_text[idx])]
        included = set(_expand_page_context(high_conf_pages, len(pages_text), radius=1))
        included.update(_collect_adjacent_detected_pages(detected, pages_text, high_conf_pages + rate_pages))
        included = sorted(included)
        schedule_text = "\n\n".join(pages_text[i] for i in included)
        return schedule_text, included
    # Use cluster-based ranges: only include pages within each tight cluster.
    # A gap of 5+ pages between detected pages = separate sections (likely spec pages in between).
    clusters = _cluster_pages(detected, max_gap=4)
    included = []
    for cluster in clusters:
        first, last = cluster[0], cluster[-1]
        included.extend(range(first, last + 1))
    included = sorted(set(included))
    schedule_text = "\n\n".join(pages_text[i] for i in included)
    return schedule_text, included


def call_claude_with_retry(
    client: anthropic.Anthropic,
    schedule_text: str,
    chunk_label: str = "",
    extra_instruction: str = "",
) -> list:
    instruction = extra_instruction or (
        "You are a Canadian construction quantity takeoff specialist. "
        "Extract EVERY item from the Schedule of Prices below. "
        "For each item return a JSON array with objects containing: "
        "item_no, spec_ref, description, quantity (number or null), unit, "
        "is_provisional (boolean), confidence (0.0-1.0). "
        "IMPORTANT: Also extract labour rates, equipment rates, hourly rates, and day-work schedule items. "
        "These are commonly in a separate 'Schedule of Additional Unit Prices' section with columns like "
        "Description, Hourly Rate, Hours, Subtotal. Extract these as items with unit='HOURS' and the hours value as quantity. "
        "If a labour/equipment rate row has no printed hours quantity, keep the row anyway with quantity=null and unit='HOURS'. "
        "If a labour/equipment rate table prints numeric row numbers, use item numbers 'L1', 'L2', ... for LABOUR rows "
        "and 'E6', 'E7', ... for EQUIPMENT rows using the printed row number. Preserve any skipped numbers exactly "
        "(for example, if the equipment list jumps from 8 to 10, there is no E9). "
        "Only if a labour table has named rows but no printed item numbers should you assign synthetic identifiers "
        "such as 'LABOUR-1', 'LABOUR-2' in order of appearance. "
        "MULTI-SECTION TENDERS: "
        "Many Ontario municipal tenders have multiple schedule sections (Part A, Part B, Parts A-C, Parts D-F, etc.) "
        "covering different streets or project areas, all within the same document. "
        "These sections use the SAME item numbers but with DIFFERENT quantities. "
        "When you detect multiple schedule sections: "
        "1. Assign a suffix letter to each section (A, B, C, D, E, F). "
        "2. Append the suffix to every item number from that section (item 1 in Part A -> '1A', item 1 in Part D -> '1D'). "
        "3. Extract ALL items from ALL sections — do NOT skip or merge items with the same number from different sections. "
        "4. The item_no field should include the section suffix: '1A', '11B', '31C', '1D', '11E', '31F'. "
        "A tender with Parts A-F and 63 items per section should produce approximately 6 x 63 = 378 item rows "
        "(minus any items not present in some sections). "
        "CRITICAL: Never merge or deduplicate items across sections. "
        "Item 11 in Part B (14,000 m\u00b2) and Item 11 in Part E (1,500 m\u00b2) are TWO SEPARATE bid items "
        "for two different project areas. Both must appear in the output with suffixes 11B and 11E. "
        "If sections are labeled with names (e.g., 'Fourth Avenue'), note this in the description. "
        "QUANTITY EXTRACTION RULES: "
        "For Lump Sum (LS) items, the quantity is ALWAYS 1 and the unit is 'LS' or 'L.S.'. "
        "NEVER extract percentage values (%, percent) as quantities. "
        "Percentages in tender documents are PAYMENT SCHEDULES (how the contractor gets paid in installments), NOT bid quantities. "
        "If you see '40% upon installation, 50% prorated, 10% upon removal' — this is a payment schedule. The quantity is still 1 LS. "
        "SPECIAL CASE: If the quantity column shows '100%' for a Lump Sum item, "
        "the correct extraction is: quantity = 1, unit = 'L.S.'. "
        "'100%' means 'one complete lump sum' — it is NOT a quantity of 100. "
        "Never extract '%' as a unit. Always convert to L.S. with quantity 1. "
        "The quantity comes ONLY from the Schedule of Prices / Bill of Quantities table columns. "
        "Do not extract quantities from Special Provisions, Payment sections, or Measurement for Payment sections. "
        "EXTRACTION SOURCE RULES: "
        "ONLY extract items from the SCHEDULE OF PRICES / FORM OF TENDER section. "
        "Do NOT extract approximate quantities from the project description, scope summary, or overview sections. "
        "Valid bid items have: an item number, a unit, and a quantity in a structured table format. "
        "Structured labour/equipment rate rows are also valid schedule items even if the quantity cell is blank. "
        "Text like 'Approximately X,XXX tonnes of...' is a summary, not a bid item. "
        "If an item has no item number, it is NOT a schedule item. Do not include it. "
        "Only include items with clear item numbers (numeric or alphanumeric like P1, LABOUR-1, E6, 1A, 11B). "
        "Return ONLY valid JSON — no markdown, no backticks, no explanation."
    )
    def _build_prompt(retry_note: str = "") -> str:
        if retry_note:
            return f"{instruction}\n\n{retry_note}\n\nSCHEDULE TEXT:\n{schedule_text}"
        return f"{instruction}\n\nSCHEDULE TEXT:\n{schedule_text}"

    retry_note = ""
    for attempt in range(1, 4):
        label = f"Extracting{' ' + chunk_label if chunk_label else ''} — attempt {attempt}/3..."
        prompt = _build_prompt(retry_note if attempt > 1 else "")
        with st.spinner(label):
            message = client.messages.create(
                model=CLAUDE_MODEL, max_tokens=16000, temperature=0,
                messages=[{"role": "user", "content": prompt}],
            )
        raw = message.content[0].text.strip()
        s, e = raw.find("["), raw.rfind("]") + 1
        if s != -1 and e > s:
            try:
                parsed_items = json.loads(raw[s:e])
                if not extra_instruction:
                    expected_sections = _expected_multi_section_letters(schedule_text)
                    if len(expected_sections) >= 3:
                        extracted_sections = _extracted_multi_section_letters(parsed_items)
                        missing_sections = sorted(expected_sections - extracted_sections)
                        if missing_sections:
                            if attempt == 3:
                                st.warning(
                                    "Extraction may be incomplete: tender contains "
                                    f"Parts {', '.join(missing_sections)} but they were not found in the parsed output."
                                )
                            else:
                                st.warning(
                                    "Extraction may be incomplete: tender contains "
                                    f"Parts {', '.join(missing_sections)} but they were not found in the parsed output. Retrying..."
                                )
                                retry_note = (
                                    "Your previous response omitted one or more schedule sections. "
                                    f"The schedule text includes Parts {', '.join(sorted(expected_sections))}. "
                                    f"Your last output missed Parts {', '.join(missing_sections)}. "
                                    "Re-read the FULL schedule and return a COMPLETE JSON array including ALL parts and suffixes. "
                                    "Return ONLY valid JSON."
                                )
                                time.sleep(2)
                                continue
                return parsed_items
            except json.JSONDecodeError as err:
                if attempt == 3:
                    st.error(f"All 3 attempts failed. Last error: {err}")
                    st.text(raw[:3000])
                    return []
                retry_note = (
                    "Your previous response was not valid JSON. "
                    "Return ONLY one COMPLETE JSON array, no markdown, no backticks, no explanation."
                )
        else:
            if attempt == 3:
                st.error("All 3 attempts failed — no JSON array found.")
                st.text(raw[:3000])
                return []
            retry_note = (
                "Your previous response did not contain a JSON array. "
                "Return ONLY one COMPLETE JSON array, no markdown, no backticks, no explanation."
            )
        time.sleep(2)
    return []


def extract_in_chunks(client: anthropic.Anthropic, schedule_text: str) -> list:
    chunks, start = [], 0
    while start < len(schedule_text):
        chunks.append(schedule_text[start : start + CHUNK_SIZE])
        start += CHUNK_SIZE - CHUNK_OVERLAP
    all_items, seen_keys = [], set()
    for i, chunk in enumerate(chunks):
        label = f"chunk {i + 1} of {len(chunks)}"
        for item in call_claude_with_retry(client, chunk, chunk_label=label):
            # Section-aware dedup: preserve suffixes and differing quantities.
            key = dedup_key(item)
            if key not in seen_keys:
                seen_keys.add(key)
                all_items.append(item)
    return all_items


def _item_no_numeric_value(item_no: object) -> float | None:
    raw = _normalize_item_no(item_no).rstrip(")")
    if not raw:
        return None
    match = re.fullmatch(r"(\d+(?:\.\d+)?)([A-Za-z])?", raw)
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def _normalize_item_no_for_scan(value: object) -> str:
    raw = _normalize_item_no(value).rstrip(")")
    if not raw:
        return ""
    match = re.fullmatch(r"(\d+(?:\.\d+)?)([A-Za-z])?", raw)
    if not match:
        return ""
    number, suffix = match.groups()
    numeric_value = _item_no_numeric_value(number)
    if numeric_value is None or not (1 <= numeric_value <= 200):
        return ""
    if not suffix:
        return number
    if suffix.islower():
        return f"{number}{suffix.lower()}"
    if suffix.upper() in "ABCDEF":
        return number
    return ""


def _page_has_recovery_schedule_signals(text: str) -> bool:
    lower = text.lower()
    marker_hits = sum(
        1
        for marker in (
            "item no", "spec. no", "spec no", "description", "tender quantity",
            "quantity", "unit", "unit price", "amount", "schedule of items and prices",
            "schedule of prices", "provisional", "labour and equipment rates",
            "schedule of additional unit prices",
        )
        if marker in lower
    )
    return marker_hits >= 3


def _extract_candidate_item_nos_from_page(text: str) -> set[str]:
    if not _page_has_recovery_schedule_signals(text):
        return set()
    pattern = re.compile(r"(?:^|\n)\s*(\d{1,3}(?:\.\d{1,2})?(?:\s*[a-z]\))?)\s+\S", re.MULTILINE)
    found = set()
    for match in pattern.finditer(text):
        normalized = _normalize_item_no_for_scan(match.group(1))
        if normalized:
            found.add(normalized)
    return found


def second_pass_extraction(
    client: anthropic.Anthropic,
    pages_text: list,
    existing_items: list,
    schedule_page_indices: list,
) -> list:
    extracted_item_nos = {
        normalized
        for normalized in (_normalize_item_no_for_scan(i.get("item_no")) for i in existing_items)
        if normalized
    }
    covered = set(schedule_page_indices)
    suspected_pages = []
    for i, page in enumerate(pages_text):
        if i in covered:
            continue
        page_item_nos = _extract_candidate_item_nos_from_page(page)
        text_only_nos = page_item_nos - extracted_item_nos
        if text_only_nos:
            suspected_pages.append(i)
    if not suspected_pages:
        return []
    suspected_pages = sorted(set(suspected_pages))
    missed_text = "\n\n".join(pages_text[i] for i in suspected_pages)
    if not missed_text.strip():
        return []
    instruction = (
        "Here are pages that may contain additional tender items not captured in the first pass. "
        "Extract ANY tender items you find. Return ONLY a JSON array with: "
        "item_no, spec_ref, description, quantity (number or null), unit, "
        "is_provisional (bool), confidence (0.0-1.0). If no items, return []."
    )
    with st.spinner(f"Second pass — {len(suspected_pages)} suspected missed page(s)..."):
        new_items = call_claude_with_retry(
            client, missed_text[:MAX_SCHEDULE_CHARS],
            chunk_label="second pass", extra_instruction=instruction,
        )
    existing_keys  = {dedup_key(i) for i in existing_items}
    existing_descs = {str(i.get("description") or "").lower().strip() for i in existing_items}
    added = []
    for item in new_items:
        key  = dedup_key(item)
        desc = str(item.get("description") or "").lower().strip()
        if key not in existing_keys and desc not in existing_descs:
            added.append(item)
    return added


def verify_extraction(items: list, full_text: str) -> list[dict]:
    results = []
    main_items = [i for i in items if _item_no_numeric_value(i.get("item_no")) is not None]
    item_nums = []
    for item in main_items:
        numeric_value = _item_no_numeric_value(item.get("item_no"))
        if numeric_value is not None:
            item_nums.append(numeric_value)
    item_nums = sorted(set(item_nums))
    gaps = []
    for i in range(len(item_nums) - 1):
        if item_nums[i + 1] - item_nums[i] > 1.5:
            gaps.append(f"{item_nums[i]:.0f}→{item_nums[i+1]:.0f}")
    if gaps:
        results.append({"check": "Item Number Sequence", "passed": False,
                        "message": f"Gaps: {', '.join(gaps)} — possible missed items"})
    else:
        results.append({"check": "Item Number Sequence", "passed": True,
                        "message": f"No gaps in {len(item_nums)} item numbers"})

    qty_pattern = re.compile(
        r"(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(m3|m2|m²|m³|ea|LS|tonnes|HOURS|ha|km)\b",
        re.IGNORECASE,
    )
    text_quantities = {m.group(1).replace(",", "").split(".")[0] for m in qty_pattern.finditer(full_text)}
    extracted_quantities = {str(int(float(i["quantity"]))) for i in items if i.get("quantity") is not None}
    missed_qty = {q for q in text_quantities - extracted_quantities if int(q) > 0}
    if len(missed_qty) > 5:
        results.append({"check": "Text Quantity Scan", "passed": False,
                        "message": f"{len(missed_qty)} quantities in document not in extraction"})
    else:
        results.append({"check": "Text Quantity Scan", "passed": True,
                        "message": f"{len(missed_qty)} minor discrepancies (likely prose text)"})

    has_total = any(p in full_text.lower() for p in
                    ("total tender price", "total bid price", "grand total", "total lump sum"))
    results.append({"check": "Total Tender Price", "passed": True,
                    "message": "Total price line found — verify all items captured"
                    if has_total else "No total price line found (may be normal)"})

    sub_items: dict = {}
    sub_pat = re.compile(r"^(\d+\.\d+)\s+([a-z])\)", re.IGNORECASE)
    for item in items:
        m = sub_pat.match(str(item.get("item_no", "")))
        if m:
            sub_items.setdefault(m.group(1), []).append(m.group(2).lower())
    broken = []
    for parent, letters in sub_items.items():
        ls = sorted(set(letters))
        if ls != [chr(ord("a") + i) for i in range(len(ls))]:
            broken.append(parent)
    if broken:
        results.append({"check": "Sub-Item Completeness", "passed": False,
                        "message": f"Gaps in sub-items: {', '.join(broken[:5])}"})
    else:
        results.append({"check": "Sub-Item Completeness", "passed": True,
                        "message": f"{len(sub_items)} sub-item groups complete"})

    count_match = re.search(r"items?\s+(\d+)\s+to\s+(\d+)", full_text.lower())
    if count_match:
        expected = int(count_match.group(2)) - int(count_match.group(1)) + 1
        if len(main_items) < expected:
            results.append({"check": "Document Item Count", "passed": False,
                            "message": f"Doc refs {expected} items but only {len(main_items)} extracted"})
        else:
            results.append({"check": "Document Item Count", "passed": True,
                            "message": f"{len(main_items)} items ≥ doc reference of {expected}"})
    else:
        results.append({"check": "Document Item Count", "passed": True,
                        "message": f"{len(main_items)} main items (no explicit count in doc)"})

    has_prov_text = "provisional" in full_text.lower() or "contingency" in full_text.lower()
    extracted_prov = sum(1 for i in items if i.get("is_provisional"))
    if has_prov_text and extracted_prov == 0:
        results.append({"check": "Provisional Items", "passed": False,
                        "message": "Doc mentions 'provisional' but no provisional items extracted"})
    else:
        results.append({"check": "Provisional Items", "passed": True,
                        "message": f"{extracted_prov} provisional items extracted"})
    return results


# ─────────────────────────────────────────────
# New helper functions — Phase 0 upgrades
# ─────────────────────────────────────────────

def analyze_cost_risks(items: list) -> list:
    """Upgrade 3: flag cost risks without any Claude call."""
    risks = []
    for item in items:
        desc = item.get("description", "").lower()
        ino  = item.get("item_no", "?")

        if any(p in desc for p in ["as directed", "as required", "to be determined",
                                    "as needed", "tbd"]):
            risks.append({
                "item": ino, "severity": "HIGH",
                "risk": "Vague scope — 'as directed/required' language",
                "advice": "Get written clarification from engineer before bidding. Price conservatively.",
            })

        if item.get("is_provisional"):
            risks.append({
                "item": ino, "severity": "MEDIUM",
                "risk": "Provisional item — may not be built",
                "advice": "Don't count on this revenue. Price at full rate but exclude from cash flow projections.",
            })

        unit = str(item.get("unit") or "").strip().upper()
        if unit in ("LS", "LUMP SUM", "LUMP") and item.get("quantity") is None:
            risks.append({
                "item": ino, "severity": "MEDIUM",
                "risk": "Lump sum item — no quantity breakdown",
                "advice": "Break down into sub-components. Risk of underbidding.",
            })

        qty = item.get("quantity")
        if qty and float(qty) > 50_000:
            risks.append({
                "item": ino, "severity": "MEDIUM",
                "risk": f"Large quantity ({float(qty):,.0f} {item.get('unit','')}) — verify against drawings",
                "advice": "Cross-check drawings. Quantity errors on large items have biggest dollar impact.",
            })

        if "contingency" in desc:
            risks.append({
                "item": ino, "severity": "LOW",
                "risk": "Contingency allowance — fixed owner amount",
                "advice": "Owner's contingency, not yours. Do not include in your cost estimate.",
            })

    seen = set()
    unique_risks = []
    for risk in risks:
        key = f"{risk['item']}_{risk['risk']}"
        if key not in seen:
            seen.add(key)
            unique_risks.append(risk)
    return unique_risks


def extract_tender_header(full_text: str) -> dict:
    """Upgrade 6: extract project metadata via regex. Fast, no API call."""
    header = {
        "project": "", "owner": "", "contract": "",
        "engineer": "", "location": "", "closing": "",
    }
    text = full_text[:HEADER_CHARS]

    # Contract number
    m = re.search(r"Contract\s+No\.?\s*:?\s*([A-Za-z0-9\-/]+)", text, re.IGNORECASE)
    if m:
        header["contract"] = m.group(1).strip()

    # Owner / municipality
    for pat in [
        r"(?:Owner|Municipality|Authority|City of|Town of|County of|Township of|Region of)\s*:?\s*([^\n]{4,60})",
        r"(Essex.Windsor\s+\w[\w\s]+Authority)",
        r"(City of [A-Z][a-z]+)",
        r"(Town of [A-Z][a-z]+)",
        r"(County of [A-Z][a-z]+)",
        r"(Region of [A-Z][a-z]+)",
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            header["owner"] = m.group(1).strip()[:80]
            break

    # Engineer / consultant
    for pat in [
        r"(?:Prepared by|Consultant|Engineer|WSP|Arcadis|CIMA|Dillon|Stantec|Jacobs)\s*[:\-]?\s*([^\n]{4,60})",
        r"(WSP\s+Canada[^\n]*)",
        r"(Stantec[^\n]*)",
        r"(CIMA[^\n]*)",
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()[:80]
            if len(val) > 3:
                header["engineer"] = val
                break

    # Tender closing date/time
    for pat in [
        r"(?:Tender Closing|Closing Date|Tenders close|Closes)\s*[:\-]?\s*([^\n]{5,60})",
        r"(?:Due|Submit|Submission)\s+(?:by|before|on)\s+([A-Z][a-z]+ \d{1,2},?\s+\d{4}[^\n]{0,30})",
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            header["closing"] = m.group(1).strip()[:80]
            break

    # Project name — first meaningful title line
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    for line in lines[:30]:
        if len(line) > 15 and re.search(r"(contract|project|construction|rehabilitation|reconstruction)", line, re.IGNORECASE):
            if not any(c in line for c in ["©", "http", "www", "@"]):
                header["project"] = line[:120]
                break

    # Location fallback
    for pat in [
        r"(?:Location|Site|Project Site)\s*:?\s*([^\n]{5,60})",
        r"(Cell \d+ [A-Za-z]+)",
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            header["location"] = m.group(1).strip()[:80]
            break

    return header


def call_claude_for_checklist(client: anthropic.Anthropic, checklist_source_text: str) -> list:
    """Upgrade 1: extract bid submission requirements."""
    prompt = (
        "You are a Canadian construction tender compliance specialist. "
        "Read this tender document and extract EVERY submission requirement the contractor must meet "
        "to submit a valid bid. Return a JSON array where each object has: "
        '{"requirement": "...", "category": one of ["Form","Insurance","Bonding","WSIB",'
        '"Certificate","Schedule","Document","Submission Requirement","Other"], '
        '"page_reference": "page X or null", "deadline": "date/timing or null", '
        '"critical": true if missing this disqualifies the bid}. '
        "Extract: bid bond, insurance certificates, WSIB clearance, agreement to bond, "
        "addenda acknowledgment, tender deposit, tender closing date/time, mandatory site meeting, "
        "HST registration, required forms. "
        "CUSTOM SUBMISSION REQUIREMENTS: scan specification and scope sections for non-standard bid-package "
        "requirements such as proposed methods, work plans, alternate-product submissions, or approaches "
        "that the bidder must include with the tender. Look for phrases like 'shall submit', "
        "'must be submitted with the bid', 'submit this proposed method as part of the tender', "
        "'provide with the tender', or 'attach to the tender'. "
        "Label these as category 'Submission Requirement' and mark them critical when the tender says "
        "they must be part of the submission. "
        "Return ONLY valid JSON array — no markdown, no backticks.\n\n"
        f"TENDER DOCUMENT:\n{checklist_source_text}"
    )
    with st.spinner("Extracting bid submission checklist..."):
        try:
            msg = client.messages.create(
                model=CLAUDE_MODEL, max_tokens=4000,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = msg.content[0].text.strip()
            s, e = raw.find("["), raw.rfind("]") + 1
            if s != -1 and e > s:
                return json.loads(raw[s:e])
        except Exception as ex:
            traceback.print_exc(file=sys.stderr)
            st.warning(f"Checklist extraction failed: {ex}")
    return []


def call_claude_for_timeline(client: anthropic.Anthropic, front_matter: str) -> list:
    """Upgrade 2: extract all dates and schedule requirements."""
    prompt = (
        "You are a Canadian construction scheduling specialist. "
        "Read this tender document and extract ALL dates, deadlines, and schedule requirements. "
        "Return a JSON array where each object has: "
        '{"event": "...", "date": "date string or null", "working_days": number or null, '
        '"flag": one of ["DEADLINE","MILESTONE","PENALTY","MEETING","INFO"], '
        '"risk_note": "any risk or concern"}. '
        "Extract: tender closing date, mandatory site meeting, contract start, completion deadline "
        "(working days), liquidated damages per day, milestone dates, holdback release, "
        "warranty period, maintenance period. "
        "Return ONLY valid JSON array — no markdown, no backticks.\n\n"
        f"TENDER DOCUMENT:\n{front_matter}"
    )
    with st.spinner("Extracting timeline & schedule requirements..."):
        try:
            msg = client.messages.create(
                model=CLAUDE_MODEL, max_tokens=3000,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = msg.content[0].text.strip()
            s, e = raw.find("["), raw.rfind("]") + 1
            if s != -1 and e > s:
                return json.loads(raw[s:e])
        except Exception as ex:
            traceback.print_exc(file=sys.stderr)
            st.warning(f"Timeline extraction failed: {ex}")
    return []


def call_claude_for_opss_full_scan(client: anthropic.Anthropic, full_text: str) -> list:
    """FIX 3: Scan the ENTIRE document for all OPSS spec references, not just schedule items."""
    # Build a comprehensive OPSS knowledge base prompt
    opss_knowledge = "\n".join(f"OPSS {code}: {desc}" for code, desc in OPSS_NOTES.items())
    prompt = (
        "You are a Canadian construction specification expert. "
        "Scan the ENTIRE tender document below for ALL OPSS spec references. "
        "Look for:\n"
        "1. Tables or lists titled 'Ontario Provincial Standard Specifications' or 'OPSS' listing applicable specs\n"
        "2. Any section saying 'The following OPSS apply to this contract' or similar\n"
        "3. OPSS numbers referenced in Special Provisions (e.g., 'OPSS.MUNI 182 and OPSS.MUNI 805 govern')\n"
        "4. OPSS numbers referenced in General Conditions sections\n"
        "5. Spec reference columns in the Schedule of Prices\n\n"
        "Return a JSON array where each unique OPSS code found is an object with: "
        '{"code": "e.g. 706", "date": "April 2018 or null", "description": "one-line description"}.\n'
        "Use these known descriptions where available (fill in if document doesn't provide one):\n"
        f"{opss_knowledge}\n\n"
        "IMPORTANT: Return each OPSS code only ONCE (deduplicate). "
        "Return ONLY valid JSON array — no markdown, no backticks.\n\n"
        f"TENDER DOCUMENT (first 80,000 chars):\n{full_text[:80000]}"
    )
    with st.spinner("Scanning full document for OPSS references..."):
        try:
            msg = client.messages.create(
                model=CLAUDE_MODEL, max_tokens=4000,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = msg.content[0].text.strip()
            s, e = raw.find("["), raw.rfind("]") + 1
            if s != -1 and e > s:
                found = json.loads(raw[s:e])
                # Dedup by code, keep first occurrence
                seen_codes = set()
                deduped = []
                for entry in found:
                    code = str(entry.get("code", "")).strip().lstrip("0")
                    if code and code not in seen_codes:
                        seen_codes.add(code)
                        entry["code"] = code
                        deduped.append(entry)
                return deduped
        except Exception as ex:
            traceback.print_exc(file=sys.stderr)
            st.warning(f"OPSS full scan failed: {ex}")
    return []


def detect_project_type(full_text: str, items: list) -> str:
    """FIX 4+5: Classify project type from tender text and extracted items."""
    text_lower = full_text.lower()[:50000]
    all_desc = " ".join(str(i.get("description", "")).lower() for i in items)

    # High-confidence override for service / maintenance contracts. This must
    # run before score-based detection because Roads Division tenders can
    # mention "road" without being road reconstruction projects.
    maintenance_primary = (
        "lifting and levelling",
        "lifting and leveling",
        "crack sealing",
        "mudjacking",
        "mud jacking",
        "mud-jacking",
        "line painting",
        "pothole patching",
        "pothole repair",
        "street sweeping",
        "grass cutting",
    )
    if any(keyword in text_lower for keyword in maintenance_primary):
        return "MAINTENANCE"

    maintenance_secondary = [
        "vendor of record",
        "no guarantee of the value or volume",
        "option to renew",
        "actual measured",
        "measured in the field",
        "estimated quantity",
    ]
    secondary_hits = sum(1 for keyword in maintenance_secondary if keyword in text_lower)
    if "maintenance" in text_lower and "maintenance of traffic" not in text_lower:
        secondary_hits += 1
    surface_terms = ("sidewalk", "patching", "sweeping", "curb", "pothole", "line painting")
    surface_hit = any(term in text_lower or term in all_desc for term in surface_terms)
    low_item_count = 0 < len(items) <= 5
    if surface_hit and (secondary_hits >= 3 or (secondary_hits >= 2 and low_item_count)):
        return "MAINTENANCE"

    # Score each project type based on keyword frequency
    scores = {
        "BRIDGE_REHAB": 0,
        "BRIDGE_REPLACEMENT": 0,
        "SEWER_WATERMAIN": 0,
        "MAINTENANCE": 0,
        "ROAD_RECONSTRUCTION": 0,
        "LANDFILL": 0,
        "CULVERT": 0,
        "ELECTRICAL_TRAFFIC": 0,
        "ENVIRONMENTAL": 0,
        "GENERAL_MUNICIPAL": 0,
        "MTO_HIGHWAY": 0,
    }

    # Bridge signals
    for kw in ["bridge rehabilitation", "repointing", "masonry", "parapet", "expansion joint",
               "bearing", "bridge deck", "abutment", "wing wall", "navigabl", "transport canada",
               "rvca", "cataraqui", "conservation authority", "dfo", "heritage"]:
        if kw in text_lower:
            scores["BRIDGE_REHAB"] += 2
    for kw in ["bridge replacement", "new bridge", "superstructure replacement"]:
        if kw in text_lower:
            scores["BRIDGE_REPLACEMENT"] += 3

    # Sewer/watermain signals
    for kw in ["sanitary sewer", "storm sewer", "watermain", "water main", "service connection",
               "bypass pump", "cathodic protection", "eca", "environmental compliance approval"]:
        if kw in text_lower or kw in all_desc:
            scores["SEWER_WATERMAIN"] += 2

    # Road reconstruction signals
    for kw in ["road reconstruction", "roadway", "asphalt", "granular", "curb and gutter",
               "sidewalk", "boulevard", "boulevard restoration"]:
        if kw in text_lower or kw in all_desc:
            scores["ROAD_RECONSTRUCTION"] += 1

    # Landfill signals
    for kw in ["landfill", "cell", "leachate", "liner", "geomembrane", "capping",
               "clay cap", "import clay", "waste management"]:
        if kw in text_lower or kw in all_desc:
            scores["LANDFILL"] += 3

    # Culvert signals
    for kw in ["culvert replacement", "culvert rehabilitation", "pipe culvert"]:
        if kw in text_lower:
            scores["CULVERT"] += 3

    # Electrical/traffic signals
    for kw in ["traffic signal", "illumination", "electrical system", "conduit"]:
        if kw in text_lower or kw in all_desc:
            scores["ELECTRICAL_TRAFFIC"] += 2

    # MTO/Highway signals
    for kw in ["ministry of transportation", "mto", "provincial highway", "400-series"]:
        if kw in text_lower:
            scores["MTO_HIGHWAY"] += 4

    # Environmental signals
    for kw in ["remediation", "contaminated soil", "environmental assessment", "phase ii esa"]:
        if kw in text_lower:
            scores["ENVIRONMENTAL"] += 4

    best = max(scores.items(), key=lambda x: x[1])
    if best[1] == 0:
        return "GENERAL_MUNICIPAL"
    return best[0]


def _parse_tender_date(value: object) -> date | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    raw = raw.split("@", 1)[0].strip()
    raw = re.sub(r"\s+", " ", raw)
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _find_timeline_date(timeline_items: list, keywords: tuple[str, ...]) -> date | None:
    for item in timeline_items:
        haystack = " ".join(
            str(item.get(field) or "").lower()
            for field in ("event", "risk_note", "date")
        )
        if any(keyword in haystack for keyword in keywords):
            parsed = _parse_tender_date(item.get("date"))
            if parsed is not None:
                return parsed
    return None


def _count_weekdays_inclusive(start: date, end: date) -> int:
    if end < start:
        return 0
    days = 0
    current = start
    while current <= end:
        if current.weekday() < 5:
            days += 1
        current += timedelta(days=1)
    return days


def _extract_bridge_work_window(timeline_items: list) -> dict | None:
    start_date = _find_timeline_date(timeline_items, ("in-water", "shoreline"))
    end_date = _find_timeline_date(timeline_items, ("completion", "final completion"))
    if start_date is None or end_date is None or end_date < start_date:
        return None
    calendar_days = (end_date - start_date).days + 1
    working_days = _count_weekdays_inclusive(start_date, end_date)
    return {
        "start_date": start_date,
        "end_date": end_date,
        "calendar_days": calendar_days,
        "working_days": working_days,
    }


def _extract_relevant_schedule_working_days(timeline_items: list) -> int | None:
    """
    Pull schedule-compression candidates from timeline rows that actually
    describe the contract work period or a completion deadline, not admin
    milestones like award, insurance, or certificate turnaround.
    """
    candidate_keywords = (
        "completion",
        "substantial completion",
        "final completion",
        "must be completed",
        "work window",
        "construction period",
        "construction duration",
        "duration",
        "working days to complete",
        "calendar days to complete",
        "contract completion",
    )
    excluded_keywords = (
        "award anticipated",
        "questions/inquiries",
        "insurance evidence",
        "certificates required",
        "warranty",
        "contract term",
        "option to renew",
        "irrevocable",
    )
    candidates = []
    for item in timeline_items:
        wd = item.get("working_days")
        if not isinstance(wd, (int, float)) or wd <= 0:
            continue
        haystack = " ".join(
            str(item.get(field) or "").lower()
            for field in ("event", "risk_note", "date")
        )
        if any(keyword in haystack for keyword in excluded_keywords):
            continue
        if any(keyword in haystack for keyword in candidate_keywords):
            candidates.append(int(wd))
    return min(candidates) if candidates else None


def generate_project_type_risks(project_type: str, full_text: str, items: list,
                                  timeline_items: list) -> list:
    """FIX 4+5: Generate project-type-aware risk flags."""
    risks = []
    text_lower = full_text.lower()
    all_desc = " ".join(str(i.get("description", "")).lower() for i in items)
    prov_count = sum(1 for i in items if i.get("is_provisional"))
    total_count = len(items)

    def _has_item_keywords(*keywords: str) -> bool:
        return any(keyword in all_desc for keyword in keywords)

    def _append_scope_risk_or_info(
        label: str,
        item_keywords: tuple[str, ...],
        bundled_keywords: tuple[str, ...],
        severity: str,
        risk_text: str,
        advice: str,
    ) -> None:
        if _has_item_keywords(*item_keywords):
            return
        is_bundled, explanation = check_if_item_is_bundled(bundled_keywords, full_text)
        if is_bundled:
            risks.append({
                "item": "INFO",
                "severity": "INFO",
                "risk": f"{label}: no separate line item — {explanation}",
                "advice": "Tender text addresses this scope outside the pricing schedule.",
            })
        else:
            risks.append({
                "item": "MISSING",
                "severity": severity,
                "risk": risk_text,
                "advice": advice,
            })

    bridge_window = None
    if project_type in ("BRIDGE_REHAB", "BRIDGE_REPLACEMENT"):
        bridge_window = _extract_bridge_work_window(timeline_items)

    # Extract working days from timeline
    working_days = None
    if bridge_window is not None:
        working_days = bridge_window["working_days"]
    else:
        working_days = _extract_relevant_schedule_working_days(timeline_items)

    # Extract liquidated damages from timeline
    ld_per_day = None
    for t in timeline_items:
        risk_note = str(t.get("risk_note") or "").lower()
        event = str(t.get("event") or "").lower()
        if "liquidated" in event or "liquidated" in risk_note:
            import re as _re
            m = _re.search(r"\$([\d,]+)", t.get("event", "") + t.get("risk_note", ""))
            if m:
                try:
                    ld_per_day = float(m.group(1).replace(",", ""))
                except Exception:
                    pass

    # ── Universal checks ──────────────────────────────────────────────────────
    if bridge_window is not None:
        start_label = bridge_window["start_date"].strftime("%B %d, %Y").replace(" 0", " ")
        end_label = bridge_window["end_date"].strftime("%B %d, %Y").replace(" 0", " ")
        risks.append({
            "item": "CONTRACT", "severity": "HIGH",
            "risk": (
                f"Compressed in-water work window: {bridge_window['calendar_days']} calendar days "
                f"({bridge_window['working_days']} working days) from {start_label} to {end_label}"
            ),
            "advice": "Front-load labour, materials, and access works. Any weather or permit delay will hit the critical path.",
        })
    elif working_days is not None and working_days < 60:
        risks.append({
            "item": "CONTRACT", "severity": "HIGH",
            "risk": f"Compressed schedule: only {working_days} working days",
            "advice": "Mobilize immediately upon award. Pre-order long-lead materials before contract execution.",
        })

    if ld_per_day is not None and ld_per_day > 500:
        risks.append({
            "item": "CONTRACT", "severity": "HIGH",
            "risk": f"High liquidated damages: ${ld_per_day:,.0f}/day",
            "advice": "Build schedule contingency. Identify critical path items now. Consider weather risk.",
        })

    if total_count > 0 and prov_count / total_count > 0.30:
        risks.append({
            "item": "CONTRACT", "severity": "MEDIUM",
            "risk": f"High provisional item ratio: {prov_count}/{total_count} items ({100*prov_count//total_count}%) provisional",
            "advice": "Revenue from provisional items not guaranteed. Build overhead recovery into firm items.",
        })

    # Check for multiple agency permits
    permit_agencies = []
    for agency_kw in ["conservation authority", "cataraqui", "rvca", "mvca", "grca",
                       "dfo", "fisheries", "transport canada", "ministry of environment",
                       "mecp", "federal", "transport canada"]:
        if agency_kw in text_lower:
            permit_agencies.append(agency_kw)
    if len(set(permit_agencies)) >= 2:
        risks.append({
            "item": "PERMITS", "severity": "HIGH",
            "risk": f"Multiple permit agencies detected: {', '.join(list(set(permit_agencies))[:3])}",
            "advice": "Coordinate permit timing early. Contractor-borne delay risk if permits delayed.",
        })

    # ── Bridge Rehab specific checks ──────────────────────────────────────────
    if project_type in ("BRIDGE_REHAB", "BRIDGE_REPLACEMENT"):
        _append_scope_risk_or_info(
            "Traffic control",
            ("traffic control", "traffic management", "tcp"),
            ("traffic control", "traffic management", "signs", "barricades", "delineators"),
            "HIGH",
            "No traffic control plan item found — required for bridge work",
            "Verify traffic control provisions in Special Conditions before bidding.",
        )
        _append_scope_risk_or_info(
            "Environmental protection",
            ("erosion", "silt", "sediment", "environmental"),
            ("erosion", "sediment", "silt", "environmental protection"),
            "MEDIUM",
            "No erosion control / environmental protection items found",
            "Check if environmental protection is included in a lump sum or provisional item.",
        )
        if not any(kw in all_desc for kw in ["scaffold", "work platform", "access"]):
            risks.append({
                "item": "MISSING", "severity": "MEDIUM",
                "risk": "No access/scaffolding provisions found — may be required for bridge work",
                "advice": "Confirm access methodology. OPSS 928 requirements apply if scaffolding used.",
            })
        if any(kw in text_lower for kw in ["heritage", "historic", "conservation authority approval"]):
            risks.append({
                "item": "HERITAGE", "severity": "MEDIUM",
                "risk": "Heritage/conservation requirements detected",
                "advice": "Heritage material matching requirements can significantly affect cost. Clarify with engineer.",
            })
        if any(kw in text_lower for kw in ["navigabl", "vessel", "shipping channel", "transport canada"]):
            risks.append({
                "item": "NAVIGABILITY", "severity": "MEDIUM",
                "risk": "Navigable waterway — Transport Canada permit likely required",
                "advice": "Transport Canada Navigable Waters Protection Act approval adds timeline risk.",
            })
        if bridge_window is None and any(kw in text_lower for kw in ["in-water work window", "fish habitat", "spawning", "dfo",
                                             "in-water work", "work window"]):
            risks.append({
                "item": "IN-WATER", "severity": "HIGH",
                "risk": "In-water work window restriction detected — compressed effective work period",
                "advice": "Identify exact window dates. Any weather delay inside this window is critical path.",
            })
        if any(kw in text_lower for kw in ["designated substance", "asbestos", "lead paint", "pcb"]):
            risks.append({
                "item": "DESIGNATED_SUBSTANCES", "severity": "MEDIUM",
                "risk": "Designated substances on site (asbestos/lead/PCB)",
                "advice": "Specialized sub-contractor required. Add handling, disposal, and air monitoring costs.",
            })
        if "p.eng" in text_lower or "professional engineer" in text_lower:
            risks.append({
                "item": "ENGINEERING", "severity": "MEDIUM",
                "risk": "P.Eng.-sealed submissions required",
                "advice": "Budget for engineering submissions: formwork, shoring, scaffolding, temporary works design.",
            })
        if "no weekend" in text_lower or "no work on weekend" in text_lower:
            risks.append({
                "item": "SCHEDULE", "severity": "MEDIUM",
                "risk": "Restrictions on weekend work detected",
                "advice": "Account for no-weekend constraint in schedule. Increases effective working days needed.",
            })
        # Measurement reclassification rules
        if any(kw in text_lower for kw in ["reclassif", "depth threshold", "pay item", "measurement rule"]):
            risks.append({
                "item": "MEASUREMENT", "severity": "MEDIUM",
                "risk": "Measurement reclassification rules in contract — depth/quantity thresholds shift work between pay items",
                "advice": "Read measurement provisions carefully. Misclassified work = no payment. Clarify before bidding.",
            })

    # ── Sewer/Watermain specific checks ──────────────────────────────────────
    elif project_type == "SEWER_WATERMAIN":
        if "dewater" not in all_desc and "dewater" not in text_lower[:20000]:
            is_bundled, explanation = check_if_item_is_bundled(("dewatering", "dewater"), full_text)
            if is_bundled:
                risks.append({
                    "item": "INFO", "severity": "INFO",
                    "risk": f"Dewatering: no separate line item — {explanation}",
                    "advice": "Tender text addresses this scope outside the pricing schedule.",
                })
            else:
                risks.append({
                    "item": "MISSING", "severity": "HIGH",
                    "risk": "No dewatering items found — high risk if high water table",
                    "advice": "Confirm groundwater conditions. Unpriced dewatering = significant cost risk.",
                })
        _append_scope_risk_or_info(
            "Erosion control",
            ("erosion", "silt", "sediment"),
            ("erosion", "sediment", "silt"),
            "MEDIUM",
            "No erosion control items found",
            "Confirm if erosion control is included in trench restoration or a separate item.",
        )
        _append_scope_risk_or_info(
            "Traffic control",
            ("traffic control", "traffic management"),
            ("traffic control", "traffic management", "signs", "barricades", "delineators"),
            "HIGH",
            "No traffic control items found — required for sewer/watermain work",
            "Verify traffic control requirements with municipality. TCP approval timeline.",
        )
        _append_scope_risk_or_info(
            "Site restoration",
            ("restoration", "trench restoration", "topsoil", "seed"),
            ("restoration", "topsoil", "seed", "boulevard"),
            "MEDIUM",
            "No trench restoration items found",
            "Confirm trench restoration scope — asphalt cut, backfill, surface restoration.",
        )
        if "bypass" not in all_desc and "bypass pump" not in text_lower[:30000]:
            if any(kw in all_desc for kw in ["sanitary sewer", "sanitary pipe", "maintenance hole"]):
                risks.append({
                    "item": "MISSING", "severity": "MEDIUM",
                    "risk": "No bypass pumping items found for sanitary work",
                    "advice": "Sanitary sewer work requires bypass pumping. Confirm if included in LS or missing.",
                })
        if "eca" in text_lower or "environmental compliance" in text_lower:
            risks.append({
                "item": "ECA", "severity": "HIGH",
                "risk": "ECA (Environmental Compliance Approval) required before work",
                "advice": "Contractor cannot commence work until ECA obtained. Add to schedule critical path.",
            })
        if any(kw in text_lower for kw in ["depth exceeds 4", "depth > 4", "deeper than 4", "> 4.5 m"]):
            risks.append({
                "item": "DEPTH", "severity": "MEDIUM",
                "risk": "Deep excavation detected (>4m) — shoring/sheeting required",
                "advice": "Deep excavation requires P.Eng. shoring design and specialized equipment.",
            })

    # ── Maintenance contract specific checks ──────────────────────────────────
    elif project_type == "MAINTENANCE":
        weather_excerpt = _extract_sentence_like_excerpt(
            full_text, r"suspended during periods of rain.{0,120}below\s*2"
        )
        if weather_excerpt:
            risks.append({
                "item": "WEATHER", "severity": "HIGH",
                "risk": f"Weather restrictions apply: {weather_excerpt}",
                "advice": "Outdoor work is suspended during rain and below 2C. Price mobilization and crew utilization for a compressed warm-weather season.",
            })
        else:
            risks.append({
                "item": "WEATHER", "severity": "INFO",
                "risk": "No explicit weather shutdown clause detected",
                "advice": "Confirm whether seasonal or temperature restrictions apply before pricing field operations.",
            })

        if re.search(r"estimates? only|no guarantee of the value or volume|subject to vary", text_lower):
            risks.append({
                "item": "VOLUME", "severity": "HIGH",
                "risk": "Quantities are estimates only — no guaranteed volume",
                "advice": "Build fixed-cost recovery into your unit rate so mobilization, insurance, and equipment costs are covered even if actual quantities are lower than estimated.",
            })

        liability_excerpt = _extract_sentence_like_excerpt(
            full_text,
            r"responsibility of the bidder|breakage due to over lifting|damage deemed excessive",
        )
        if liability_excerpt:
            risks.append({
                "item": "LIABILITY", "severity": "MEDIUM",
                "risk": f"Bidder is liable for breakage and workmanship defects: {liability_excerpt}",
                "advice": "Inspect existing conditions carefully and carry contingency for slab damage, rework, and replacement exposure.",
            })

        if re.search(r"cash discount.{0,120}(award|consideration|taken into consideration)", text_lower, re.DOTALL):
            risks.append({
                "item": "AWARD", "severity": "MEDIUM",
                "risk": "Cash discount is part of award evaluation",
                "advice": "Prompt-payment discount terms can influence award. Consider whether a competitive discount strengthens the bid without harming cash flow.",
            })

        irrevocable_days = _parse_irrevocable_days(full_text, timeline_items)
        if irrevocable_days and irrevocable_days > 90:
            risks.append({
                "item": "IRREVOCABLE", "severity": "MEDIUM",
                "risk": f"Tender is irrevocable for {irrevocable_days} days ({round(irrevocable_days / 30, 1):g} months)",
                "advice": "Hold pricing for the full irrevocable period and account for seasonal labour and material availability changes.",
            })

        if "option to renew" in text_lower and re.search(r"prices?.{0,60}firm|firm.{0,60}first year", text_lower, re.DOTALL):
            risks.append({
                "item": "RENEWAL", "severity": "MEDIUM",
                "risk": "Contract renewal terms exist but pricing flexibility may be limited",
                "advice": "Clarify whether renewal-year pricing can be adjusted. If not, build inflation and wage escalation risk into the first-year rate.",
            })

    # ── Road Reconstruction specific checks ───────────────────────────────────
    elif project_type == "ROAD_RECONSTRUCTION":
        _append_scope_risk_or_info(
            "Erosion control",
            ("erosion", "silt", "sediment"),
            ("erosion", "sediment", "silt"),
            "MEDIUM",
            "No erosion control items found",
            "Required during construction. May be included in mobilization/general conditions.",
        )
        _append_scope_risk_or_info(
            "Traffic control",
            ("traffic control", "traffic management"),
            ("traffic control", "traffic management", "signs", "barricades", "delineators"),
            "HIGH",
            "No traffic control items found",
            "Required for all road work. Confirm TCP requirements with municipal engineer.",
        )
        if "october 15" in text_lower or "october 31" in text_lower:
            risks.append({
                "item": "SEASONAL", "severity": "HIGH",
                "risk": "Paving season deadline — asphalt placement may be restricted after October",
                "advice": "Northern Ontario paving season ends mid-October. Compressed schedule risk.",
            })

    # ── Landfill specific checks ──────────────────────────────────────────────
    elif project_type == "LANDFILL":
        _append_scope_risk_or_info(
            "Erosion control",
            ("erosion", "silt", "sediment"),
            ("erosion", "sediment", "silt"),
            "MEDIUM",
            "No erosion control items found",
            "Required for landfill site work. Check if included in earthwork items.",
        )
        if not any(kw in all_desc for kw in ["leachate", "leachate pipe", "leachate collect"]):
            risks.append({
                "item": "MISSING", "severity": "MEDIUM",
                "risk": "No leachate management items found — verify scope",
                "advice": "Landfill cells typically require leachate collection. Confirm scope with engineer.",
            })
        if not any(kw in all_desc for kw in ["quality assurance", "qa", "testing", "geosynthetic"]):
            risks.append({
                "item": "QA", "severity": "MEDIUM",
                "risk": "Quality assurance/testing requirements — verify who pays for third-party QA",
                "advice": "Landfill liner QA is typically owner-engaged. Clarify before bidding.",
            })
        if not any(kw in text_lower for kw in ["environmental monitor", "ground water", "groundwater"]):
            risks.append({
                "item": "ENV_MONITORING", "severity": "MEDIUM",
                "risk": "No environmental monitoring items found",
                "advice": "Verify environmental monitoring requirements during construction.",
            })

    return risks


def build_project_type_risk_section(project_type: str, risks: list) -> list:
    """Return formatted risk entries with project type label prepended."""
    formatted = [{"item": "PROJECT_TYPE", "severity": "INFO",
                   "risk": f"Detected project type: {project_type}",
                   "advice": "Risk analysis is tailored to this project type."}]
    formatted.extend(risks)
    return formatted


def generate_missing_warnings(project_type: str, items: list, full_text: str = "") -> tuple[list, list]:
    """
    FIX 3: Generate project-type-aware missing scope warnings.
    Replaces the hardcoded 4-warning block that incorrectly fired for all project types.
    BRIDGE_REHAB: check traffic control, environmental protection, access/scaffolding only.
    SEWER_WATERMAIN: dewatering, erosion, traffic, restoration.
    MAINTENANCE: skip infrastructure warnings that do not apply to service contracts.
    ROAD_RECONSTRUCTION: traffic, erosion, restoration.
    LANDFILL: environmental, leachate, erosion.
    DEFAULT: original generic 4-warning set.
    """
    warnings = []
    info_notes = []
    all_desc = " ".join(str(it.get("description", "")).lower() for it in items)

    def _has_item_keywords(*keywords: str) -> bool:
        return any(keyword in all_desc for keyword in keywords)

    def _add_warning_or_info(
        message: str,
        item_keywords: tuple[str, ...],
        bundled_keywords: tuple[str, ...],
    ) -> None:
        if _has_item_keywords(*item_keywords):
            return
        is_bundled, explanation = check_if_item_is_bundled(bundled_keywords, full_text)
        if is_bundled:
            info_notes.append(f"{message.replace('No ', '').replace(' found', '')}: no separate line item — {explanation}")
        else:
            warnings.append(message)

    if project_type in ("BRIDGE_REHAB", "BRIDGE_REPLACEMENT"):
        # Bridge rehab: traffic control, environmental protection, access/scaffolding
        # Do NOT warn about dewatering, site restoration, granular backfill, topsoil/sod
        _add_warning_or_info(
            "No traffic control items found",
            ("traffic control", "traffic management", "tcp", "traffic"),
            ("traffic control", "traffic management", "signs", "barricades", "delineators"),
        )
        _add_warning_or_info(
            "No environmental protection / erosion control items found",
            ("erosion", "silt", "sediment", "environmental"),
            ("erosion", "silt", "sediment", "environmental protection"),
        )
        if not _has_item_keywords("scaffold", "work platform", "access to work area", "access"):
            warnings.append("No access / scaffolding provisions found (OPSS 928 may apply)")

    elif project_type == "SEWER_WATERMAIN":
        _add_warning_or_info(
            "No dewatering items found",
            ("dewater",),
            ("dewatering", "dewater"),
        )
        _add_warning_or_info(
            "No erosion control items found",
            ("erosion", "silt", "sediment"),
            ("erosion", "silt", "sediment"),
        )
        _add_warning_or_info(
            "No traffic control items found",
            ("traffic", "sign"),
            ("traffic control", "traffic management", "signs", "barricades", "delineators"),
        )
        _add_warning_or_info(
            "No site restoration items found",
            ("restoration", "topsoil", "seed"),
            ("restoration", "topsoil", "seed", "boulevard"),
        )

    elif project_type == "MAINTENANCE":
        if not _has_item_keywords("traffic", "sign", "traffic control"):
            is_bundled, explanation = check_if_item_is_bundled(
                ("traffic control", "traffic management", "signs", "barricades", "delineators"),
                full_text,
            )
            if is_bundled:
                info_notes.append(f"Traffic control: no separate line item — {explanation}")
            else:
                warnings.append("No traffic control items found")
        if re.search(r"cleaning up|leave the site.{0,40}clean|debris", full_text, re.IGNORECASE):
            info_notes.append("Site cleanup requirements are addressed in the tender specifications.")
        if re.search(r"warranty|performance surety|bond", full_text, re.IGNORECASE):
            info_notes.append("Warranty and/or surety obligations are addressed in the tender text.")

    elif project_type == "ROAD_RECONSTRUCTION":
        _add_warning_or_info(
            "No traffic control items found",
            ("traffic", "sign"),
            ("traffic control", "traffic management", "signs", "barricades", "delineators"),
        )
        _add_warning_or_info(
            "No erosion control items found",
            ("erosion", "silt", "sediment"),
            ("erosion", "silt", "sediment"),
        )
        _add_warning_or_info(
            "No site restoration items found",
            ("restoration", "topsoil", "seed"),
            ("restoration", "topsoil", "seed", "boulevard"),
        )

    elif project_type == "LANDFILL":
        _add_warning_or_info(
            "No erosion control items found",
            ("erosion", "silt", "sediment"),
            ("erosion", "silt", "sediment"),
        )
        if not _has_item_keywords("leachate", "leachate pipe"):
            warnings.append("No leachate management items found — verify scope")
        _add_warning_or_info(
            "No site restoration items found",
            ("restoration", "topsoil", "seed"),
            ("restoration", "topsoil", "seed", "boulevard"),
        )

    else:
        # DEFAULT / GENERAL_MUNICIPAL: original 4-warning set
        _add_warning_or_info(
            "No erosion control items found",
            ("erosion", "silt", "sediment"),
            ("erosion", "silt", "sediment"),
        )
        _add_warning_or_info(
            "No dewatering items found",
            ("dewater",),
            ("dewatering", "dewater"),
        )
        _add_warning_or_info(
            "No traffic control items found",
            ("traffic", "sign"),
            ("traffic control", "traffic management", "signs", "barricades", "delineators"),
        )
        _add_warning_or_info(
            "No site restoration items found",
            ("restoration", "topsoil", "seed"),
            ("restoration", "topsoil", "seed", "boulevard"),
        )

    return warnings, info_notes


def build_xlsx(
    items: list,
    opss_refs: list,
    missing_warnings: list,
    val_warnings: list,
    missing_info_notes: list | None = None,
    cost_risks: list | None = None,
    checklist_items: list | None = None,
    timeline_items: list | None = None,
    opss_notes_map: dict | None = None,
    other_standards: list | None = None,
    possible_items: list | None = None,
    summary_rows: list | None = None,
    debug_info: dict | None = None,
    project_type: str | None = None,
) -> BytesIO:
    """6-sheet workbook: Takeoff, Summary, OPSS Notes, Strategy & Risks, Bid Checklist, Timeline."""
    wb = openpyxl.Workbook()
    cost_risks      = cost_risks or []
    checklist_items = checklist_items or []
    timeline_items  = timeline_items or []
    opss_notes_map  = opss_notes_map or OPSS_NOTES
    missing_info_notes = missing_info_notes or []
    other_standards = other_standards or []
    possible_items  = possible_items or []
    summary_rows    = summary_rows or []

    # ── Sheet 1: Takeoff ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Takeoff"
    h1 = ["Item No", "Spec Ref", "Description", "Quantity", "Unit",
          "Provisional", "Confidence", "Category"]
    _write_header(ws1, h1, HEADER_FILL, HEADER_FONT)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(h1))}1"
    for r, item in enumerate(items, 2):
        conf   = item.get("confidence", 0.5)
        is_prov = bool(item.get("is_provisional"))
        ws1.cell(r, 1, item.get("item_no", ""))
        ws1.cell(r, 2, item.get("spec_ref", ""))
        ws1.cell(r, 3, item.get("description", ""))
        ws1.cell(r, 4, item.get("quantity"))
        ws1.cell(r, 5, item.get("unit", ""))
        ws1.cell(r, 6, "Yes" if is_prov else "No")
        ws1.cell(r, 7, round(conf, 2))
        ws1.cell(r, 8, item.get("category", "General"))
        fill = PROV_FILL if is_prov else (LOW_CONF_FILL if conf < 0.5 else None)
        if fill:
            for c in range(1, 9):
                ws1.cell(r, c).fill = fill

    # ── Possible Additional Items section (bottom of Takeoff sheet) ───────────
    if possible_items:
        sep_row = len(items) + 3  # blank row gap
        sep_cell = ws1.cell(sep_row, 1, f"POSSIBLE ADDITIONAL ITEMS ({len(possible_items)}) — Verify before bidding")
        sep_cell.fill = SECTION_FILL
        sep_cell.font = SECTION_FONT
        ws1.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=len(h1))
        for c, h in enumerate(h1, 1):
            cell = ws1.cell(sep_row + 1, c, h)
            cell.fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        for r, item in enumerate(possible_items, sep_row + 2):
            conf = item.get("confidence", 0.5)
            ws1.cell(r, 1, item.get("item_no", ""))
            ws1.cell(r, 2, item.get("spec_ref", ""))
            ws1.cell(r, 3, item.get("description", ""))
            ws1.cell(r, 4, item.get("quantity"))
            ws1.cell(r, 5, item.get("unit", ""))
            ws1.cell(r, 6, "Yes" if item.get("is_provisional") else "No")
            ws1.cell(r, 7, round(conf, 2))
            ws1.cell(r, 8, item.get("category", "General"))
            for c in range(1, 9):
                ws1.cell(r, c).fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    # ── Summary rows section (Tender Price, Contingency, HST, Total) ────────────
    if summary_rows:
        # Calculate the next available row in ws1
        last_data_row = ws1.max_row
        sum_sep_row = last_data_row + 2  # one blank row gap
        sep_cell = ws1.cell(sum_sep_row, 1, f"SUMMARY / TOTALS ({len(summary_rows)} rows)")
        sep_cell.fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
        sep_cell.font = Font(bold=True, color="FFFFFF", size=10)
        ws1.merge_cells(start_row=sum_sep_row, start_column=1,
                        end_row=sum_sep_row, end_column=len(h1))
        sum_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        for r, srow in enumerate(summary_rows, sum_sep_row + 1):
            ws1.cell(r, 1, srow.get("item_no", ""))
            ws1.cell(r, 2, srow.get("spec_ref", ""))
            ws1.cell(r, 3, srow.get("description", ""))
            ws1.cell(r, 4, "")
            ws1.cell(r, 5, "")
            ws1.cell(r, 6, "No")
            ws1.cell(r, 7, 1.0)
            ws1.cell(r, 8, "Summary")
            for c in range(1, 9):
                ws1.cell(r, c).fill = sum_fill

    _autosize(ws1)

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    _write_header(ws2, ["Category", "Item Count"], HEADER_FILL, HEADER_FONT)
    cat_counts: dict = {}
    unit_totals: dict = {}
    for item in items:
        cat = item.get("category", "General")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        qty  = item.get("quantity")
        unit = str(item.get("unit") or "").strip().lower()
        if qty is not None and unit not in ("missing unit", "check manually", ""):
            unit_totals.setdefault(unit, 0.0)
            unit_totals[unit] += float(qty)
    for r, (cat, cnt) in enumerate(sorted(cat_counts.items()), 2):
        ws2.cell(r, 1, cat)
        ws2.cell(r, 2, cnt)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = "A1:B1"
    offset = len(cat_counts) + 4
    _write_header(ws2, ["Unit", "Total Quantity"], HEADER_FILL, HEADER_FONT, start_row=offset)
    for r, (unit, total) in enumerate(sorted(unit_totals.items()), offset + 1):
        ws2.cell(r, 1, unit)
        ws2.cell(r, 2, round(total, 3))
    _autosize(ws2)

    # ── Sheet 3: OPSS Notes ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("OPSS Notes")
    notes_header = ["OPSS Code", "Description"]
    _write_header(ws3, notes_header, HEADER_FILL, HEADER_FONT)
    ws3.freeze_panes = "A2"
    if opss_refs:
        for r, code in enumerate(opss_refs, 2):
            ws3.cell(r, 1, f"OPSS {code}")
            ws3.cell(r, 2, opss_notes_map.get(code, "No description available"))
    else:
        ws3.cell(2, 1, "No OPSS specifications referenced in this tender.")
    _autosize(ws3)

    # ── Sheet 4: Strategy & Risks ────────────────────────────────────────────
    ws4 = wb.create_sheet("Strategy & Risks")
    row = 1

    def _section(ws, r, title):
        cell = ws.cell(r, 1, f"  {title}")
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        return r + 1

    def _sub_header(ws, r, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(r, c, h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        return r + 1

    # Section A: Cost Risk Flags
    row = _section(ws4, row, "A — COST RISK FLAGS")
    row = _sub_header(ws4, row, ["Item No", "Severity", "Risk", "Advice", ""])
    sev_fills = {"HIGH": RISK_HIGH, "MEDIUM": RISK_MED, "LOW": RISK_LOW}
    for risk in cost_risks:
        sev = risk.get("severity", "LOW")
        ws4.cell(row, 1, risk.get("item", ""))
        ws4.cell(row, 2, sev)
        ws4.cell(row, 3, risk.get("risk", ""))
        ws4.cell(row, 4, risk.get("advice", ""))
        f = sev_fills.get(sev)
        if f:
            for c in range(1, 5):
                ws4.cell(row, c).fill = f
        row += 1
    if not cost_risks:
        ws4.cell(row, 1, "No cost risks flagged"); row += 1
    row += 1

    # Section B: Missing Scope Items
    row = _section(ws4, row, "B — MISSING SCOPE ITEMS")
    row = _sub_header(ws4, row, ["Type", "Message", "", "", ""])
    for w in missing_warnings:
        ws4.cell(row, 1, "WARNING")
        ws4.cell(row, 2, w)
        row += 1
    for note in missing_info_notes:
        ws4.cell(row, 1, "INFO")
        ws4.cell(row, 2, note)
        row += 1
    if not missing_warnings and not missing_info_notes:
        ws4.cell(row, 1, "All key item categories present"); row += 1
    row += 1

    # Section C: OPSS Compliance
    section_c_title = "C — OPSS COMPLIANCE REQUIREMENTS" if opss_refs or not other_standards else "C — STANDARDS & COMPLIANCE REFERENCES"
    row = _section(ws4, row, section_c_title)
    c_header = ["OPSS Code", "Description", "", "", ""] if opss_refs or not other_standards else ["Standard", "Description", "", "", ""]
    row = _sub_header(ws4, row, c_header)
    if opss_refs:
        for code in opss_refs:
            ws4.cell(row, 1, f"OPSS {code}")
            ws4.cell(row, 2, opss_notes_map.get(code, "No description available"))
            row += 1
    elif other_standards:
        for std in other_standards:
            ws4.cell(row, 1, std.get("code", ""))
            ws4.cell(row, 2, std.get("description", ""))
            row += 1
    else:
        ws4.cell(row, 1, "No OPSS references detected"); row += 1
    row += 1

    # Section D: Bid Tips
    row = _section(ws4, row, "D — BID TIPS & INTELLIGENCE")
    prov_items   = [i for i in items if i.get("is_provisional")]
    ls_items     = [i for i in items if str(i.get("unit") or "").upper() in ("LS", "LUMP SUM", "LUMP")]
    # FIX 6: Differentiate force account / equipment rate items from true provisional items
    # Force account items: item_no like P-1, P-2, E-1, E-10 OR descriptions of labour/equipment rates
    import re as _re
    _fa_no_pat = _re.compile(r"^(?:[PE]|FA|EQ)-\d+$", _re.IGNORECASE)
    _fa_desc_kws = (
        "superintendent", "equipment operator", "labourer", "laborer",
        "compactor", "excavator", "bulldozer", "grader", "loader",
        "truck", "pump", "backhoe", "crane", "roller",
    )
    force_account_items = [
        i for i in prov_items
        if _fa_no_pat.match(str(i.get("item_no") or ""))
        or any(kw in str(i.get("description") or "").lower() for kw in _fa_desc_kws)
    ]
    true_prov_items = [i for i in prov_items if i not in force_account_items]
    # Top 5 by quantity
    qty_items = [(i, float(i["quantity"])) for i in items if i.get("quantity") is not None]
    qty_items.sort(key=lambda x: -x[1])
    # Build provisional tip with force account split
    if force_account_items:
        prov_tip = (
            f"PROVISIONAL ITEMS: {len(true_prov_items)} scope items flagged provisional. "
            f"Additionally, {len(force_account_items)} force account / equipment hourly rates are included "
            f"for potential time-and-materials work — these are rate schedule items, not scope uncertainty."
        )
    else:
        prov_tip = (
            f"PROVISIONAL ITEMS: {len(prov_items)} items flagged provisional. "
            f"Revenue not guaranteed — exclude from cash flow projections."
        )
    tips = [
        prov_tip,
        f"LUMP SUM ITEMS: {len(ls_items)} lump sum items — these are areas where contractors commonly underbid. Break each down before pricing.",
        f"HIGH RISK ITEMS: {sum(1 for r in cost_risks if r['severity']=='HIGH')} HIGH severity risks — require engineer clarification before bidding.",
    ]
    if project_type == "MAINTENANCE":
        tips.append("UNIT RATE STRATEGY: Estimated quantities and no guaranteed volume mean your unit rate must recover mobilization, insurance, and equipment costs even at reduced actual quantities.")
        tips.append("SERVICE CONTRACT: Strong performance on the initial term can support renewals. Price sustainably and treat the first term as a relationship-building contract.")
    if qty_items:
        tips.append(f"LARGEST ITEMS BY QUANTITY (cross-check against drawings):")
    for item, qty in qty_items[:5]:
        tips.append(f"  • Item {item.get('item_no','?')}: {item.get('description','')[:60]} — {qty:,.0f} {item.get('unit','')}")
    row = _sub_header(ws4, row, ["Tip", "", "", "", ""])
    for tip in tips:
        ws4.cell(row, 1, tip); row += 1
    row += 1

    # Section E: Validation Issues
    row = _section(ws4, row, "E — DATA QUALITY / VALIDATION ISSUES")
    row = _sub_header(ws4, row, ["Issue", "", "", "", ""])
    for w in val_warnings:
        ws4.cell(row, 1, w); row += 1
    if not val_warnings:
        ws4.cell(row, 1, "No validation issues"); row += 1

    ws4.column_dimensions["A"].width = 15
    ws4.column_dimensions["B"].width = 12
    ws4.column_dimensions["C"].width = 55
    ws4.column_dimensions["D"].width = 60
    ws4.freeze_panes = "A2"

    # ── Sheet 5: Bid Checklist ────────────────────────────────────────────────
    ws5 = wb.create_sheet("Bid Checklist")
    _write_header(ws5, ["✓", "Requirement", "Category", "Deadline", "Critical", "Page Ref"],
                  HEADER_FILL, HEADER_FONT)
    ws5.freeze_panes = "A2"
    ws5.auto_filter.ref = "A1:F1"
    for r, item in enumerate(checklist_items, 2):
        ws5.cell(r, 1, "☐")
        ws5.cell(r, 2, item.get("requirement", ""))
        ws5.cell(r, 3, item.get("category", ""))
        ws5.cell(r, 4, item.get("deadline") or "")
        critical = item.get("critical", False)
        ws5.cell(r, 5, "YES" if critical else "")
        ws5.cell(r, 6, item.get("page_reference") or "")
        if critical:
            for c in range(1, 7):
                ws5.cell(r, c).fill = CRITICAL_FILL
                ws5.cell(r, c).font = Font(bold=True, color="FFFFFF")
    if not checklist_items:
        ws5.cell(2, 2, "No checklist items extracted")
    ws5.column_dimensions["A"].width = 5
    ws5.column_dimensions["B"].width = 60
    ws5.column_dimensions["C"].width = 14
    ws5.column_dimensions["D"].width = 25
    ws5.column_dimensions["E"].width = 10
    ws5.column_dimensions["F"].width = 12

    # ── Sheet 6: Timeline & Schedule ─────────────────────────────────────────
    ws6 = wb.create_sheet("Timeline & Schedule")
    _write_header(ws6, ["Event", "Date", "Working Days", "Flag", "Risk Note"],
                  HEADER_FILL, HEADER_FONT)
    ws6.freeze_panes = "A2"
    ws6.auto_filter.ref = "A1:E1"
    flag_fills = {
        "DEADLINE":  PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "PENALTY":   PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "MILESTONE": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "MEETING":   PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
        "INFO":      None,
    }
    for r, item in enumerate(timeline_items, 2):
        ws6.cell(r, 1, item.get("event", ""))
        ws6.cell(r, 2, item.get("date") or "")
        wd = item.get("working_days")
        ws6.cell(r, 3, wd if wd is not None else "")
        flag = item.get("flag", "INFO")
        ws6.cell(r, 4, flag)
        ws6.cell(r, 5, item.get("risk_note") or "")
        f = flag_fills.get(flag)
        if f:
            for c in range(1, 6):
                ws6.cell(r, c).fill = f
    if not timeline_items:
        ws6.cell(2, 1, "No timeline items extracted")
    ws6.column_dimensions["A"].width = 55
    ws6.column_dimensions["B"].width = 20
    ws6.column_dimensions["C"].width = 14
    ws6.column_dimensions["D"].width = 12
    ws6.column_dimensions["E"].width = 55

    # ── Debug sheet ─────────────────────────────────────────────────────────────
    if debug_info:
        wsd = wb.create_sheet("Debug")
        wsd.column_dimensions["A"].width = 35
        wsd.column_dimensions["B"].width = 80
        debug_rows = [
            ("Code Version",                  debug_info.get("code_version", "?")),
            ("Project Type Detected",          debug_info.get("project_type", "?")),
            ("OPSS Codes Found (regex)",       ", ".join(debug_info.get("opss_regex", []))),
            ("OPSS Codes Found (API)",         ", ".join(debug_info.get("opss_api", []))),
            ("Summary Rows Found",             str(len(debug_info.get("summary_rows", [])))),
        ]
        for item in debug_info.get("schedule_items", []):
            debug_rows.append((str(item.get("description", ""))[:100], item.get("category", "?")))
        for r, (label, value) in enumerate(debug_rows, 1):
            wsd.cell(r, 1, label)
            wsd.cell(r, 2, value)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_header(ws, headers, fill, font, start_row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


# ─────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────
with st.sidebar:
    # ── Company Profile ───────────────────────
    st.subheader("Company Profile")
    if "company_profile" not in st.session_state:
        with st.form("company_profile"):
            company_name = st.text_input("Company Name")
            location = st.text_input("Location (city)")
            trades = st.multiselect(
                "Your Trades",
                ["General Contractor", "Sewer & Watermain", "Road & Paving",
                 "Structural/Concrete", "Electrical", "Earthwork",
                 "Landscaping", "Demolition", "Fencing"],
            )
            crew_size = st.selectbox("Crew Size", ["1-5", "6-15", "16-30", "31-50", "50+"])
            typical_project = st.selectbox(
                "Typical Project Size",
                ["Under $500K", "$500K - $1M", "$1M - $5M", "$5M - $10M", "$10M+"],
            )
            submitted = st.form_submit_button("Save Profile")
            if submitted and company_name:
                st.session_state.company_profile = {
                    "name": company_name,
                    "location": location,
                    "trades": trades,
                    "crew_size": crew_size,
                    "typical_project": typical_project,
                }
                st.success(f"Profile saved for {company_name}")
    else:
        profile = st.session_state.company_profile
        st.write(f"**{profile['name']}**")
        st.write(f"{profile['location']} | {profile['crew_size']} crew")
        if profile["trades"]:
            st.write(f"Trades: {', '.join(profile['trades'])}")
        if st.button("Edit Profile"):
            del st.session_state.company_profile
            st.rerun()
    st.divider()

    st.header("OPSS Intelligence")
    st.caption("Matching OPSS specs will appear here after extraction.")
    opss_placeholder = st.empty()
    st.divider()
    st.header("Cross-Verification")
    verify_placeholder = st.empty()
    st.divider()
    st.header("Missing Item Warnings")
    warnings_placeholder = st.empty()
    st.divider()
    st.header("Validation Issues")
    val_placeholder = st.empty()
    st.divider()

    # ── Scan History ──────────────────────────
    st.header("Scan History")
    if os.path.exists(HISTORY_FILE):
        try:
            with open(HISTORY_FILE) as _hf:
                _history = json.load(_hf)
            if _history:
                for _scan in reversed(_history[-10:]):
                    st.caption(
                        f"📄 {_scan.get('filename','?')} — "
                        f"{_scan.get('total_items','?')} items — "
                        f"{_scan.get('date_scanned','?')}"
                    )
            else:
                st.caption("No scans yet.")
        except Exception:
            st.caption("History unavailable.")
    else:
        st.caption("No scans yet.")

# ─────────────────────────────────────────────
# Main UI
# ─────────────────────────────────────────────
uploaded = st.file_uploader("Upload Schedule of Prices PDF", type=["pdf"])

trade_filter = st.selectbox(
    "Trade Filter",
    ["All Trades", "Sewer & Drainage", "Road & Paving", "Structural", "Electrical"],
)

full_scan = st.checkbox(
    "Extended Analysis — reads full document (additional processing time and cost)",
    value=False,
)

with st.expander("Pre-Scan Checklist", expanded=True):
    addenda_count = st.number_input("How many addenda were issued?", min_value=0, value=0)
    addenda_incorporated = st.checkbox("I have incorporated all addenda into my documents")
    if addenda_count > 0 and not addenda_incorporated:
        st.error("STOP — incorporate all addenda before scanning. Missing an addendum can disqualify your bid.")

extract_btn = st.button("Extract", type="primary", disabled=(uploaded is None))

# ─────────────────────────────────────────────
# Extraction pipeline
# ─────────────────────────────────────────────
if extract_btn and uploaded:
    t_start = time.time()
    for key in ("items", "xlsx_buffer", "val_warnings", "missing_warnings",
                "opss_refs", "stats", "verify_results",
                "cost_risks", "checklist_items", "timeline_items", "tender_header",
                "full_text", "project_type"):
        st.session_state.pop(key, None)

    chars_used = 0  # track total chars sent to Claude for cost estimate

    # Step 1: Read PDF
    with st.spinner("Reading PDF..."):
        pdf_bytes  = uploaded.read()
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        except Exception as e:
            st.error(f"Could not open this PDF: {e}. Make sure the file is not password-protected or corrupted.")
            st.stop()
        pages_text_raw = [page.get_text() for page in doc]
        num_pages  = len(pages_text_raw)

    # pdfalign preprocessing
    pages_text = []
    for i, raw_text in enumerate(pages_text_raw):
        if HAS_PDFALIGN:
            try:
                aligned = pdfalign_align(pdf_bytes, page_number=i)
                pages_text.append(aligned if aligned and aligned.strip() else raw_text)
            except Exception:
                pages_text.append(raw_text)
        else:
            pages_text.append(raw_text)

    full_text = "\n\n".join(pages_text)

    # Step 2: Build schedule text
    with st.spinner("Detecting schedule pages..."):
        schedule_text, schedule_page_indices = build_schedule_text(pages_text, full_scan)
        num_schedule_pages = len(schedule_page_indices)
        if not schedule_page_indices:
            st.warning("No schedule pages detected — scanning full document.")
        if full_scan:
            st.info(f"Deep scan: processing all {num_pages} pages.")
        else:
            span = (
                f"pages {schedule_page_indices[0]+1}–{schedule_page_indices[-1]+1}"
                if schedule_page_indices else "full document"
            )
            st.info(f"{num_schedule_pages} schedule pages ({span}) detected.")

    # Step 3: Extract items with Claude
    client = anthropic.Anthropic(api_key=_ANTHROPIC_API_KEY)
    if len(schedule_text) > MAX_SCHEDULE_CHARS:
        st.info(f"Large document ({len(schedule_text):,} chars) — splitting into chunks.")
        items_raw  = extract_in_chunks(client, schedule_text)
        chars_used += len(schedule_text)
    else:
        items_raw  = call_claude_with_retry(client, schedule_text[:MAX_SCHEDULE_CHARS])
        chars_used += min(len(schedule_text), MAX_SCHEDULE_CHARS)

    # Step 4: Second pass
    second_pass_count = 0
    if not full_scan and items_raw:
        new_items = second_pass_extraction(client, pages_text, items_raw, schedule_page_indices)
        if new_items:
            second_pass_count = len(new_items)
            st.success(f"Second pass found {second_pass_count} additional item(s).")
            items_raw.extend(new_items)

    # Step 5: Validate + split by quality
    with st.spinner("Validating extraction..."):
        filtered_items, filter_warnings = filter_non_schedule_items(items_raw)
        filtered_items, hst_warnings = fix_hst_percentage_items(filtered_items)
        filtered_items, lump_sum_warnings = fix_lump_sum_quantities(filtered_items)
        filtered_items, labour_warnings = ensure_labour_rate_items(schedule_text, filtered_items)
        filtered_items, numbered_rate_warnings = ensure_numbered_rate_items(schedule_text, filtered_items)
        filtered_items, force_account_warnings = ensure_force_account_rate_items(schedule_text, filtered_items, full_text=full_text)
        for item in filtered_items:
            item["category"] = categorize_item(item.get("description", ""), item.get("unit", ""))
            print(f"[DEBUG CAT] item_no={item.get('item_no','?')!r:8} cat={item.get('category','?'):20} desc={str(item.get('description',''))[:80]!r}", file=sys.stderr, flush=True)
        all_validated, val_warnings = validate_extraction(filtered_items)
        val_warnings = (
            filter_warnings
            + hst_warnings
            + lump_sum_warnings
            + labour_warnings
            + numbered_rate_warnings
            + force_account_warnings
            + val_warnings
        )
        items, possible_items = split_items_by_quality(all_validated)
        if possible_items:
            st.info(
                f"Quality filter: {len(items)} confirmed items → Takeoff sheet. "
                f"{len(possible_items)} unconfirmed items → 'Possible Additional Items' section."
            )

    # FIX 4: Extract summary/total rows (Contingency, HST, Total Tender Price)
    print(f"[DEBUG SUMMARY] schedule_text length={len(schedule_text)} chars", file=sys.stderr, flush=True)
    summary_rows = extract_summary_rows(schedule_text)
    print(f"[DEBUG SUMMARY] matched {len(summary_rows)} summary rows: {[r.get('description','?') for r in summary_rows]}", file=sys.stderr, flush=True)

    # Step 6: OPSS refs — explicit matches gate any OPSS enrichment
    opss_refs_from_items = extract_opss_refs(items)
    print(f"[DEBUG OPSS] full_text length={len(full_text)} chars", file=sys.stderr, flush=True)
    opss_regex_codes = extract_opss_from_full_text(full_text)
    print(f"[DEBUG OPSS] regex pass found: {opss_regex_codes}", file=sys.stderr, flush=True)
    print(f"[DEBUG OPSS] pass1 (from items): {opss_refs_from_items}", file=sys.stderr, flush=True)
    explicit_opss_codes = sorted(
        set(opss_refs_from_items + opss_regex_codes),
        key=lambda x: int(x) if x.isdigit() else 9999,
    )

    opss_full_scan_results = []
    opss_full_scan_codes = []
    if explicit_opss_codes:
        opss_full_scan_results = call_claude_for_opss_full_scan(client, full_text)
        chars_used += min(len(full_text), 80000)
        opss_full_scan_codes = [str(e.get("code", "")).strip() for e in opss_full_scan_results if e.get("code")]
        print(f"[DEBUG OPSS] claude pass found: {opss_full_scan_codes}", file=sys.stderr, flush=True)
        opss_refs = sorted(
            set(explicit_opss_codes + opss_full_scan_codes),
            key=lambda x: int(x) if x.isdigit() else 9999,
        )
    else:
        print(
            "[DEBUG OPSS] hallucination guard active: no explicit OPSS refs in items/full text, skipping Claude/DB OPSS enrichment",
            file=sys.stderr,
            flush=True,
        )
        opss_refs = []

    other_standards = extract_other_standards_from_full_text(full_text) if not opss_refs else []
    # Build enriched note map: prefer full scan descriptions over hardcoded ones
    opss_note_map = {}
    for entry in opss_full_scan_results:
        code = str(entry.get("code", "")).strip()
        if code:
            desc = entry.get("description") or OPSS_NOTES.get(code, f"OPSS {code} — see spec document")
            if entry.get("date"):
                desc = f"{desc} [{entry['date']}]"
            opss_note_map[code] = desc
    # Fill remaining codes with DB / hardcoded fallback
    db_map = get_opss_notes_from_db([c for c in opss_refs if c not in opss_note_map]) if opss_refs else {}
    for code in opss_refs:
        if code not in opss_note_map:
            opss_note_map[code] = db_map.get(code, OPSS_NOTES.get(code, f"OPSS {code} — see spec document"))

    # Step 7: Project type detection + project-type-aware missing scope warnings (FIX 3)
    # Detect project type early so warnings are tailored to scope (not generic for all tenders)
    project_type = detect_project_type(full_text, items)
    print(f"[DEBUG PROJTYPE] detected project_type={project_type!r}", file=sys.stderr, flush=True)
    missing_warnings, missing_info_notes = generate_missing_warnings(project_type, items, full_text=full_text)
    print(
        f"[DEBUG PROJTYPE] generate_missing_warnings returned {len(missing_warnings)} warnings and "
        f"{len(missing_info_notes)} info notes: warnings={missing_warnings} info={missing_info_notes}",
        file=sys.stderr,
        flush=True,
    )

    # Step 8: Cross-verification
    with st.spinner("Running cross-verification..."):
        verify_results = verify_extraction(items, full_text)

    # Step 9: Tender header (regex, no extra API call)
    tender_header = extract_tender_header(full_text)

    # Step 10: Bid submission checklist (Claude call on front matter)
    checklist_source_text = build_checklist_source_text(full_text)
    checklist_items = call_claude_for_checklist(client, checklist_source_text)
    chars_used += len(checklist_source_text)

    # Step 11: Timeline extraction (Claude call on front matter)
    front_matter = full_text[:FRONT_MATTER_CHARS]
    timeline_items = call_claude_for_timeline(client, front_matter)
    chars_used += len(front_matter)

    # Step 12: Cost risk analysis (Python only)
    cost_risks = analyze_cost_risks(items)

    # FIX 3+4+5: Project type already detected in Step 7; use it for project-type-aware risk engine
    # project_type is set above (Step 7); timeline_items now available for risk scoring
    project_type_risks = generate_project_type_risks(project_type, full_text, items, timeline_items)
    print(f"[DEBUG PROJTYPE] generate_project_type_risks returned {len(project_type_risks)} risks: {[r.get('risk','?') for r in project_type_risks]}", file=sys.stderr, flush=True)
    # Prepend project-type risks to cost_risks
    cost_risks = build_project_type_risk_section(project_type, project_type_risks) + cost_risks

    # Step 13: Build XLSX (6 sheets)
    with st.spinner("Building spreadsheet..."):
        xlsx_buffer = build_xlsx(
            items, opss_refs, missing_warnings, val_warnings,
            missing_info_notes=missing_info_notes,
            cost_risks=cost_risks, checklist_items=checklist_items, timeline_items=timeline_items,
            opss_notes_map=opss_note_map,
            other_standards=other_standards,
            possible_items=possible_items,
            summary_rows=summary_rows,
            project_type=project_type,
            debug_info={
                "code_version":  "f2ea006",
                "project_type":  project_type,
                "opss_regex":    opss_regex_codes,
                "opss_api":      opss_full_scan_codes,
                "summary_rows":  summary_rows,
                "schedule_items": [
                    {"description": it.get("description", ""), "category": it.get("category", "?")}
                    for it in items
                ],
            },
        )

    t_elapsed      = time.time() - t_start
    api_cost       = (chars_used / CHARS_PER_TOKEN / 1_000_000) * COST_PER_M_INPUT

    # Build a brief tender summary string for Q&A context
    cat_counts_summary: dict = {}
    for it in items:
        c = it.get("category", "General")
        cat_counts_summary[c] = cat_counts_summary.get(c, 0) + 1
    breakdown_str = ", ".join(f"{c}: {n}" for c, n in sorted(cat_counts_summary.items(), key=lambda x: -x[1]))
    tender_summary = (
        f"Tender: {tender_header.get('project', uploaded.name)} | "
        f"Owner: {tender_header.get('owner', 'Unknown')} | "
        f"Contract: {tender_header.get('contract', 'N/A')} | "
        f"Total items: {len(items)} | "
        f"Categories: {breakdown_str}"
    )

    st.session_state.update({
        "items":             items,
        "xlsx_buffer":       xlsx_buffer,
        "val_warnings":      val_warnings,
        "missing_warnings":  missing_warnings,
        "missing_info_notes": missing_info_notes,
        "opss_refs":         opss_refs,
        "opss_note_map":     opss_note_map,
        "other_standards":   other_standards,
        "verify_results":    verify_results,
        "cost_risks":        cost_risks,
        "checklist_items":   checklist_items,
        "timeline_items":    timeline_items,
        "tender_header":     tender_header,
        "tender_summary":    tender_summary,
        "project_type":      project_type,
        "full_text":         full_text,   # FIX 1: store for Q&A
        "extraction_done":   True,
        "messages":          [],
        "question_count":    0,
        "stats": {
            "elapsed":        t_elapsed,
            "pages":          num_pages,
            "schedule_pages": num_schedule_pages,
            "chars_sent":     chars_used,
            "api_cost":       api_cost,
            "second_pass":    second_pass_count,
            "full_scan":      full_scan,
        },
    })

    # Step 14: Save tender history
    try:
        history = []
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE) as hf:
                history = json.load(hf)
        history.append({
            "tender_id":         str(uuid.uuid4()),
            "filename":          uploaded.name,
            "date_scanned":      date.today().isoformat(),
            "total_items":       len(items),
            "total_value_estimate": None,
            "company":           st.session_state.get("company_profile", {}).get("name", ""),
            "project_type":      (sorted(cat_counts_summary.items(), key=lambda x: -x[1])[0][0]
                                  if cat_counts_summary else "unknown"),
            "municipality":      tender_header.get("owner", ""),
            "outcome":           None,
        })
        with open(HISTORY_FILE, "w") as hf:
            json.dump(history, hf, indent=2)
    except Exception:
        pass  # History save is non-critical

# ─────────────────────────────────────────────
# Results (from session state)
# ─────────────────────────────────────────────
items = st.session_state.get("items")
if items:
    xlsx_buffer     = st.session_state["xlsx_buffer"]
    val_warnings    = st.session_state["val_warnings"]
    missing_warnings = st.session_state["missing_warnings"]
    opss_refs       = st.session_state["opss_refs"]
    verify_results  = st.session_state.get("verify_results", [])
    cost_risks      = st.session_state.get("cost_risks", [])
    checklist_items = st.session_state.get("checklist_items", [])
    timeline_items  = st.session_state.get("timeline_items", [])
    tender_header   = st.session_state.get("tender_header", {})
    stats           = st.session_state.get("stats", {})

    # ── Upgrade 6: Tender Summary Header ─────────────────────────────────────
    if any(tender_header.values()):
        with st.container(border=True):
            h1, h2, h3 = st.columns(3)
            with h1:
                if tender_header.get("project"):
                    st.markdown(f"**PROJECT**  \n{tender_header['project']}")
                if tender_header.get("owner"):
                    st.markdown(f"**OWNER**  \n{tender_header['owner']}")
            with h2:
                if tender_header.get("contract"):
                    st.markdown(f"**CONTRACT**  \n{tender_header['contract']}")
                if tender_header.get("engineer"):
                    st.markdown(f"**ENGINEER**  \n{tender_header['engineer']}")
            with h3:
                if tender_header.get("location"):
                    st.markdown(f"**LOCATION**  \n{tender_header['location']}")
                if tender_header.get("closing"):
                    st.markdown(f"**TENDER CLOSING**  \n:red[{tender_header['closing']}]")

    # ── Metrics ───────────────────────────────────────────────────────────────
    total          = len(items)
    with_qty       = sum(1 for it in items if it.get("quantity") is not None)
    lump_sum_cnt   = sum(1 for it in items if str(it.get("unit", "")).upper() in ("LS", "LUMP SUM", "LUMP"))
    provisional_cnt = sum(1 for it in items if it.get("is_provisional"))
    avg_conf       = sum(it.get("confidence", 0.5) for it in items) / total if total else 0
    checks_passed  = sum(1 for v in verify_results if v["passed"]) if verify_results else 0
    checks_total   = len(verify_results)

    st.divider()
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Total Items", total)
    c2.metric("With Quantities", with_qty)
    c3.metric("Lump Sum", lump_sum_cnt)
    c4.metric("Provisional", provisional_cnt)
    c5.metric("Confidence Score", f"{avg_conf:.2f}")
    c6.metric("Verification", f"{checks_passed}/{checks_total} ✓")

    # Trade breakdown
    cat_counts: dict = {}
    for it in items:
        cat = it.get("category", "General")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    breakdown = ", ".join(
        f"{cat}: {cnt}" for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1])
    )
    st.caption(f"Trade breakdown — {breakdown}")

    # ── Trade filter & table ──────────────────────────────────────────────────
    df = pd.DataFrame(items)
    trade_kw_map = {
        "Sewer & Drainage": ["sewer", "drain", "storm", "sanitary", "culvert", "manhole", "catch basin"],
        "Road & Paving":    ["asphalt", "granular", "paving", "boulevard", "curb", "sidewalk", "road", "grading"],
        "Structural":       ["concrete", "footing", "structure", "bridge", "retaining", "reinforc"],
        "Electrical":       ["electrical", "conduit", "wire", "light", "signal", "cabinet"],
    }
    if trade_filter != "All Trades":
        kws  = trade_kw_map.get(trade_filter, [])
        mask = df["description"].str.lower().apply(lambda d: any(k in str(d) for k in kws))
        df_show = df[mask]
    else:
        df_show = df
    st.dataframe(df_show, use_container_width=True)

    # ── Downloads ─────────────────────────────────────────────────────────────
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        st.download_button(
            "Download Excel (6 sheets)", data=xlsx_buffer,
            file_name="mestre_takeoff.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col_dl2:
        st.download_button(
            "Download JSON", data=json.dumps(items, indent=2),
            file_name="mestre_takeoff.json", mime="application/json",
        )

    # ── Upgrade 1: Bid Submission Checklist ───────────────────────────────────
    with st.expander(f"📋 Bid Submission Checklist ({len(checklist_items)} requirements)", expanded=True):
        if checklist_items:
            critical = [i for i in checklist_items if i.get("critical")]
            if critical:
                st.error(f"⚠️ {len(critical)} CRITICAL requirements — missing any of these disqualifies your bid")

            # Group by category
            by_cat: dict = {}
            for item in checklist_items:
                cat = item.get("category", "Other")
                by_cat.setdefault(cat, []).append(item)

            for cat in CHECKLIST_CATEGORIES:
                cat_items = by_cat.get(cat, [])
                if not cat_items:
                    continue
                st.markdown(f"**{cat}**")
                for j, req in enumerate(cat_items):
                    label = req.get("requirement", "")
                    if req.get("deadline"):
                        label += f"  *(due: {req['deadline']})*"
                    if req.get("page_reference"):
                        label += f"  *(p. {req['page_reference']})*"
                    prefix = "🔴 " if req.get("critical") else ""
                    st.checkbox(f"{prefix}{label}", key=f"chk_{cat}_{j}")
        else:
            st.info("No checklist items extracted. The front matter may be non-standard or missing.")

    # ── Upgrade 2: Timeline & Schedule ────────────────────────────────────────
    with st.expander(f"📅 Timeline & Schedule ({len(timeline_items)} items)"):
        if timeline_items:
            flag_colors = {
                "DEADLINE":  "🔴", "PENALTY": "🛑",
                "MILESTONE": "🟢", "MEETING": "🔵", "INFO": "⚪",
            }
            tl_rows = []
            for item in timeline_items:
                flag = item.get("flag", "INFO")
                tl_rows.append({
                    "Flag":         f"{flag_colors.get(flag, '⚪')} {flag}",
                    "Event":        item.get("event", ""),
                    "Date":         item.get("date") or "—",
                    "Working Days": item.get("working_days") or "—",
                    "Risk Note":    item.get("risk_note") or "",
                })
            tl_df = pd.DataFrame(tl_rows)
            st.dataframe(tl_df, use_container_width=True, hide_index=True)
        else:
            st.info("No timeline items extracted.")

    # ── Upgrade 3: Cost Risk Indicators ──────────────────────────────────────
    high_risks = [r for r in cost_risks if r["severity"] == "HIGH"]
    med_risks  = [r for r in cost_risks if r["severity"] == "MEDIUM"]
    low_risks  = [r for r in cost_risks if r["severity"] == "LOW"]

    with st.expander(
        f"⚠️ Cost Risk Indicators — "
        f"{len(high_risks)} HIGH · {len(med_risks)} MEDIUM · {len(low_risks)} LOW"
    ):
        if cost_risks:
            if high_risks:
                st.markdown("#### 🔴 HIGH SEVERITY")
                for r in high_risks:
                    st.error(f"**Item {r['item']}** — {r['risk']}  \n💡 {r['advice']}")
            if med_risks:
                st.markdown("#### 🟠 MEDIUM SEVERITY")
                for r in med_risks:
                    st.warning(f"**Item {r['item']}** — {r['risk']}  \n💡 {r['advice']}")
            if low_risks:
                st.markdown("#### 🟡 LOW SEVERITY")
                for r in low_risks:
                    st.info(f"**Item {r['item']}** — {r['risk']}  \n💡 {r['advice']}")
        else:
            st.success("No cost risks flagged.")

    # ── Quick Cost Estimate ────────────────────────────────────────────────────
    with st.expander("Quick Cost Estimate"):
        st.caption("Enter unit prices to estimate total cost.")
        cost_rows = []
        for it in items:
            qty = it.get("quantity")
            if qty is not None:
                cost_rows.append({
                    "Item No": it.get("item_no", ""),
                    "Description": it.get("description", "")[:60],
                    "Quantity": qty,
                    "Unit": it.get("unit", ""),
                    "Unit Price ($)": 0.0,
                    "Total ($)": 0.0,
                })
        if cost_rows:
            cost_df = pd.DataFrame(cost_rows)
            edited  = st.data_editor(
                cost_df,
                column_config={
                    "Unit Price ($)": st.column_config.NumberColumn(min_value=0, format="$%.2f"),
                    "Total ($)":      st.column_config.NumberColumn(format="$%.2f"),
                },
                disabled=["Item No", "Description", "Quantity", "Unit", "Total ($)"],
                use_container_width=True, key="cost_editor",
            )
            edited["Total ($)"] = edited["Quantity"] * edited["Unit Price ($)"]
            st.metric("Estimated Grand Total", f"${edited['Total ($)'].sum():,.2f}")
        else:
            st.info("No items with quantities found for cost estimation.")

    # ── Upgrade 7: Engine Stats ───────────────────────────────────────────────
    with st.expander("Engine Stats"):
        if stats:
            s1, s2, s3, s4, s5 = st.columns(5)
            s1.metric("Extraction Time",   f"{stats['elapsed']:.1f}s")
            s2.metric("Pages Processed",   stats["pages"])
            s3.metric("Pages Analyzed",    stats["schedule_pages"])
            s4.metric("Input Volume",      f"{stats['chars_sent']:,}")
            s5.metric("Engine Units",      f"{stats['api_cost'] * 10:.1f} EU")
            if stats.get("second_pass"):
                st.caption(f"Second pass recovered {stats['second_pass']} additional item(s).")
            if stats.get("full_scan"):
                st.caption("Deep scan mode was used.")
            st.divider()
            v1, v2, v3 = st.columns(3)
            v1.metric("Engine Units",       f"{stats['api_cost'] * 10:.1f} EU")
            v2.info("💼 **Contractor value**  \nReplaces ~4–8 hours of manual takeoff work")
            v3.success("💰 **MESTRE price**  \n$29 per scan")

    # ── Sidebar: OPSS Intelligence ────────────────────────────────────────────
    _opss_note_map = st.session_state.get("opss_note_map", OPSS_NOTES)
    _other_standards = st.session_state.get("other_standards", [])
    with opss_placeholder.container():
        if opss_refs:
            for code in opss_refs:
                st.markdown(f"**OPSS {code}**")
                st.caption(_opss_note_map.get(code, "No description available"))
        elif _other_standards:
            for std in _other_standards:
                st.markdown(f"**{std.get('code', '')}**")
                st.caption(std.get("description", ""))
        else:
            st.caption("No matching OPSS or other standards references found.")

    # ── Sidebar: Cross-Verification ───────────────────────────────────────────
    with verify_placeholder.container():
        if verify_results:
            passed = sum(1 for v in verify_results if v["passed"])
            total_checks = len(verify_results)
            if passed == total_checks:
                st.success(f"All {total_checks} checks passed")
            elif passed >= total_checks - 1:
                st.warning(f"{passed}/{total_checks} checks passed")
            else:
                st.error(f"{passed}/{total_checks} checks passed")
            for v in verify_results:
                if v["passed"]:
                    st.markdown(f"✅ **{v['check']}**")
                    st.caption(v["message"])
                else:
                    st.markdown(f"🔴 **{v['check']}**")
                    st.warning(v["message"])
        else:
            st.caption("Verification results appear here after extraction.")

    # ── Sidebar: Missing Item Warnings ────────────────────────────────────────
    with warnings_placeholder.container():
        if missing_warnings:
            for w in missing_warnings:
                st.warning(w)
        else:
            st.success("All key item categories present.")

    # ── Sidebar: Validation Issues ────────────────────────────────────────────
    with val_placeholder.container():
        if val_warnings:
            for w in val_warnings:
                st.warning(w)
        else:
            st.success("No validation issues.")

    # ── Trade mismatch warning (company profile) ──────────────────────────────
    profile = st.session_state.get("company_profile")
    if profile and profile.get("trades"):
        contractor_trades = [t.lower() for t in profile["trades"]]
        # Check for significant electrical scope
        electrical_items = [
            it for it in items
            if any(kw in str(it.get("description", "")).lower()
                   for kw in ["electrical", "conduit", "wiring", "lighting", "signal"])
        ]
        has_electrical_scope = len(electrical_items) >= 3
        contractor_does_electrical = any("electrical" in t for t in contractor_trades)
        if has_electrical_scope and not contractor_does_electrical:
            item_refs = ", ".join(
                f"Item {it['item_no']}" for it in electrical_items[:5]
            )
            st.warning(
                f"This tender includes significant electrical scope "
                f"({item_refs}{'...' if len(electrical_items) > 5 else ''}) — "
                f"consider subcontracting if not in your capabilities."
            )

elif not extract_btn:
    st.info("Upload a PDF and click Extract to begin.")

# ─────────────────────────────────────────────
# Q&A Chat Interface
# ─────────────────────────────────────────────
def count_questions(text: str) -> int:
    """Estimate distinct user questions conservatively for the quota warning."""
    cleaned = re.sub(r'"[^"]*\?"', '"QUOTED"', text)
    cleaned = re.sub(r"'[^']*\?'", "'QUOTED'", cleaned)

    explicit_questions = re.findall(r"([^?]+\?)", cleaned, re.DOTALL)
    follow_up_patterns = (
        "what's the rule",
        "what is the rule",
        "what's the detail",
        "what are the details",
        "what's the process",
        "what is the process",
        "can you clarify",
        "how so",
        "why is that",
    )

    explicit_count = 0
    for q in explicit_questions:
        q_clean = " ".join(q.strip().lower().split())
        if explicit_count > 0 and any(q_clean.startswith(pattern) for pattern in follow_up_patterns):
            continue
        explicit_count += 1

    if explicit_count > 0:
        return explicit_count

    sentences = re.split(r"[.!?\n]", cleaned)
    question_words = (
        "what", "how", "why", "when", "where",
        "is ", "are ", "can ", "does ", "did ",
        "will ", "would ", "should ", "could ",
    )
    q_word_count = sum(
        1
        for sentence in sentences
        if (s := sentence.strip().lower()) and len(s) > 15 and any(s.startswith(qw) for qw in question_words)
    )
    return max(q_word_count, 1)


def get_question_count_from_response(response_text: str, fallback_count: int) -> int:
    """Parse the structured QUESTION_COUNT prefix from Sonnet's response."""
    match = re.search(r"QUESTION_COUNT:\s*(\d+)", response_text)
    if match:
        return int(match.group(1))
    match = re.search(r"I see (\d+) questions? in your message", response_text)
    if match:
        return int(match.group(1))
    return fallback_count


def clean_question_count_prefix(response_text: str) -> str:
    """Remove the QUESTION_COUNT metadata line before displaying the answer."""
    return re.sub(r"^\s*QUESTION_COUNT:\s*\d+\s*\n?", "", response_text, count=1, flags=re.MULTILINE).strip()


if st.session_state.get("extraction_done"):
    st.markdown("---")
    st.subheader("Ask about this tender")
    st.caption("5 questions included per scan. Ask about scope, requirements, risks, or strategy.")

    # Display chat history
    for msg in st.session_state.get("messages", []):
        with st.chat_message(msg["role"]):
            content = msg["content"]
            if msg["role"] == "assistant":
                content = content.replace("$", "\\$")
            st.write(content)

    if "question_count" not in st.session_state:
        st.session_state.question_count = 0

    used = st.session_state.question_count
    remaining = max(0, 5 - used)
    st.caption(f"Questions used: {used}/5 ({remaining} remaining)")

    if prompt := st.chat_input("Ask about this tender..."):
        num_questions = count_questions(prompt)
        remaining = max(0, 5 - st.session_state.question_count)

        if remaining <= 0:
            st.warning("You've used your 5 included questions. Additional questions: $2 each.")
            st.stop()

        if num_questions > remaining:
            st.warning(
                f"Your message contains {num_questions} questions but you only have "
                f"{remaining} remaining. They'll all be counted — extras at $2 each."
            )

        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.write(prompt)

        # FIX 1: Build comprehensive context from all extraction results + full tender text
        items_context = json.dumps(st.session_state.get("items", [])[:50], indent=2)

        # Build timeline text
        _tl = st.session_state.get("timeline_items", [])
        timeline_text = "\n".join(
            f"- [{i.get('flag','INFO')}] {i.get('event','')} | Date: {i.get('date') or 'N/A'} "
            f"| Working Days: {i.get('working_days') or 'N/A'} | Risk: {i.get('risk_note') or ''}"
            for i in _tl
        ) if _tl else "No timeline data extracted."

        # Build checklist text
        _cl = st.session_state.get("checklist_items", [])
        checklist_text = "\n".join(
            f"- [{i.get('category','')}] {i.get('requirement','')} | Deadline: {i.get('deadline') or 'N/A'} "
            f"| Critical: {'YES' if i.get('critical') else 'no'}"
            for i in _cl
        ) if _cl else "No checklist data extracted."

        # Build OPSS notes text
        _opss = st.session_state.get("opss_refs", [])
        _opss_map = st.session_state.get("opss_note_map", OPSS_NOTES)
        _other_standards = st.session_state.get("other_standards", [])
        if _opss:
            opss_notes_text = "\n".join(
                f"- OPSS {code}: {_opss_map.get(code, 'See spec document')}"
                for code in _opss
            )
        elif _other_standards:
            opss_notes_text = "\n".join(
                f"- {std.get('code', '')}: {std.get('description', '')}"
                for std in _other_standards
            )
        else:
            opss_notes_text = "No OPSS specs or other standards detected."

        # Build strategy/risks text
        _risks = st.session_state.get("cost_risks", [])
        strategy_risks_text = "\n".join(
            f"- [{r.get('severity','?')}] Item {r.get('item','?')}: {r.get('risk','')} | {r.get('advice','')}"
            for r in _risks
        ) if _risks else "No risk flags."

        # Full tender text (FIX 1: include full text, not just front matter)
        tender_full_text = st.session_state.get("full_text", "")
        project_type_str = st.session_state.get("project_type", "UNKNOWN")

        qa_context = f"""## EXTRACTED TAKEOFF ITEMS
{items_context}

## TIMELINE & SCHEDULE
{timeline_text}

## BID CHECKLIST & SUBMISSION REQUIREMENTS
{checklist_text}

## OPSS COMPLIANCE NOTES
{opss_notes_text}

## STRATEGY & RISK FLAGS
{strategy_risks_text}

## DETECTED PROJECT TYPE
{project_type_str}

## TENDER SUMMARY
{st.session_state.get('tender_summary', 'Not available')}

## FULL TENDER TEXT
{tender_full_text}"""

        with st.chat_message("assistant"):
            with st.spinner("MESTRE™ is analyzing your tender..."):
                _client = anthropic.Anthropic(api_key=_ANTHROPIC_API_KEY)
                response = _client.messages.create(
                    model=CLAUDE_MODEL,
                    max_tokens=3000,
                    messages=[
                        {
                            "role": "user",
                            "content": (
                                "You are MESTRE, a Canadian construction bidding intelligence assistant. "
                                "You have read and analyzed this tender document in full. "
                                "Help the contractor understand the tender and develop their bid strategy.\n\n"
                                "IMPORTANT RULES FOR ANSWERING:\n"
                                "1. ALWAYS check the Timeline & Schedule data FIRST for any question about dates, deadlines, work windows, liquidated damages, or timing.\n"
                                "2. ALWAYS check the Bid Checklist for any question about bonds, deposits, insurance, submission requirements, or tender form requirements.\n"
                                "3. ALWAYS check the Strategy & Risk Flags for any question about risks, provisional items, or missing scope.\n"
                                "4. ALWAYS search the Full Tender Text for specific contract language before saying 'check the tender documents.' If the answer is in the text, QUOTE the relevant provision and give a direct answer.\n"
                                "5. NEVER say 'check the tender documents' or 'this information isn't in the extracted data' if the answer exists ANYWHERE in the context provided. Search ALL sections before responding.\n"
                                "6. When answering questions about contract mechanisms (measurement rules, payment terms, reclassification triggers), find and cite the EXACT contract language. Do not guess or give generic advice.\n"
                                "7. For yes/no questions, give the YES or NO answer FIRST, then cite the contract provision that supports it.\n"
                                "8. Calculate derived values when possible (e.g., working days between two dates, excluding weekends and holidays).\n"
                                "9. Before answering, count the number of DISTINCT questions being asked. Count topics, not raw question marks.\n"
                                "10. A short follow-up like 'What's the rule?' or 'What are the details?' attached to the same topic is still one question.\n"
                                "11. Ignore question marks inside quoted tender text.\n"
                                "12. Start your response with exactly one line in this format: QUESTION_COUNT: [number]\n\n"
                                f"COMPREHENSIVE TENDER DATA:\n{qa_context}\n\n"
                                f"The system detected {num_questions} question(s) in this message. "
                                f"The contractor is asking: {prompt}\n\n"
                                f"If {num_questions} > 1, start your response with "
                                f"'I see {num_questions} questions in your message:' then answer each one "
                                "with a clear heading. "
                                "Answer specifically based on this tender. Be practical, direct, and reference "
                                "specific item numbers and contract clauses when relevant."
                            ),
                        }
                    ],
                )
                raw_answer = response.content[0].text
                claude_count = get_question_count_from_response(raw_answer, num_questions)
                st.session_state.question_count += claude_count
                answer = clean_question_count_prefix(raw_answer)
                # Escape dollar signs so Streamlit doesn't render as LaTeX
                display_answer = answer.replace("$", "\\$")
                st.write(display_answer)
                st.session_state.messages.append({"role": "assistant", "content": answer})

        used_now = st.session_state.question_count
        remaining_now = max(0, 5 - used_now)
        st.caption(f"Questions used: {used_now}/5 ({remaining_now} remaining)")
